import atexit
import datetime
import json
import os
import platform
import re
import shutil
import signal
import sqlite3
import subprocess
import sys
import tempfile
import threading
import time
import traceback
import tkinter as tk
from dataclasses import asdict, dataclass
from email import policy
from email.parser import BytesParser
from email.utils import parseaddr
from pathlib import Path
from typing import Any, Literal, Never, TypeAlias, get_args
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook  # type: ignore

VERSION = "0.4"

CONFIG_FILE = "robotruntime_config.json"

def load_or_create_config(path: str = CONFIG_FILE):
    @dataclass(frozen=True)
    class RuntimeConfig:

        rpa_tool_claim_timeout: int = 10        # Max wait time for RPA tool to claim workflow
        rpa_tool_execution_timeout: int = 10    # (demo friendly) max wait time for RPA tool to finish workflow
        poll_interval: int = 1                  # (demo-friendly) poll interval for runtime_loop()
        queryflow_poll_interval: int = 1         # (demo-friendly) poll interval for query_flow

        # eg 05:00 to 23:00
        operating_hours_start: int = 5
        operating_hours_end: int = 23

        system_log_path: str = "system.log"
        handover_file: str = "handover.json"
        audit_db_path: str = "job_audit.db"
        friends_path: str = "friends.xlsx"

        recordings_in_progress_folder: str = "recordings_in_progress"
        recordings_destination_folder: str = "recordings_destination" # A (demo-friendly) local destination, move to a shared drive accessabile to all users

        network_healthcheck_path: str | None = None    # None is a (demo-friendly) always healthy path, replace with eg "G:\\"
        rpa_admin_email: str = "rpa_admin@company.local"

    # below written by AI chatgpt instant 5.3
    default_config = RuntimeConfig()

    if not os.path.exists(path):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(asdict(default_config), f, indent=2)
        return default_config

    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    allowed_keys = set(RuntimeConfig.__dataclass_fields__.keys())

    unknown_keys = set(raw.keys()) - allowed_keys
    if unknown_keys:
        raise ValueError(f"Unknown config keys in {path}: {sorted(unknown_keys)}")

    return RuntimeConfig(**raw)


# ============================================================
# DATA MODELS
# ============================================================

HandoverState: TypeAlias = Literal["idle", "job_queued", "job_running", "job_verifying", "safestop"]
JobName: TypeAlias = str    # or JobName: TypeAlias = Literal["ping", "qty_adjust", "po_adjust", "order_adjust", ... ]
SourceType: TypeAlias = Literal["personal_inbox", "shared_inbox", "erp_query"]
LifecycleStatus: TypeAlias = Literal["REJECTED", "QUEUED", "RUNNING", "VERIFYING", "FAIL", "DONE"]
DashboardStatus: TypeAlias = Literal["online", "safestop", "working", "no_network" , "out_of_office"]
RuntimeErrorCode: TypeAlias = Literal["PRE_HANDOVER_CRASH", "RPA_TOOL_CRASH", "VERIFICATION_MISMATCH",  "VERIFICATION_TIMEOUT", "POST_HANDOVER_CRASH", "OUT_OF_SERVICE","OUTSIDE_WORKING_HOURS", "UNKNOWN_JOB", "NO_ACCESS", "NO_NETWORK", "INVALID_INPUT", "CODE_ERROR", "RECOVERY_SOURCE_MISSING", "IN_SAFESTOP"]


@dataclass
class HandoverFile:
    """Payload stored in handover.json and exchanged with the RPA tool."""
    
    state: HandoverState

    source_ref: str | None = None           # backend identifier eg. Outlook EntryID or "ERP_ORDER:12345" (demo uses filename for email backend)
    source_type: SourceType | None = None   # eg. "personal_inbox"
    job_name: JobName | None = None         # eg. "Ping"
    job_id: int | None = None               # eg. 202611051223

    email_address: str | None = None        # for email
    email_subject: str | None = None        # for email
    email_body: str | None = None           # for email eg. "Hi, change the order 12345 to 44 pcs"

    parsed_source_data: dict[str, Any] | None = None    # eg. {"order_number": 12345, "target_qty": 44}
    rpatool_payload: dict[str, Any] | None = None       # eg. {"order_number": 12345, "target_qty": 44, "pick_qty_from_location": "WH7",} (this is the final data sent to RPA Tool)

@dataclass
class JobCandidate:
    '''Included in HandoverFile if the candidate is accepted.'''

    source_ref: str
    source_type: SourceType
    parsed_source_data: dict[str, Any]

    email_address: str | None = None
    email_subject: str | None = None
    email_body: str | None = None

@dataclass
class QueryWorkItem:
    candidate: JobCandidate
    rpatool_payload: dict[str, Any]

@dataclass
class JobResult:
    is_success: bool
    error_message: str | None = None
    error_code: RuntimeErrorCode | None = None
    rpatool_payload: dict[str, Any] | None = None


class RuntimeFault(Exception):
    error_code: RuntimeErrorCode = "CODE_ERROR"

    def __init__(self, message:str, job_id:int|None=None, handover_file:HandoverFile|None=None, cause:Exception|None=None, traceback_text:str|None=None):
        super().__init__(message)
        self.error_message = message
        self.job_id = job_id
        self.handover_file = handover_file
        self.cause = cause
        self.traceback_text = traceback_text
        self.error_code = self.__class__.error_code
class PreHandoverCrash(RuntimeFault):
    error_code: RuntimeErrorCode = "PRE_HANDOVER_CRASH"
class RpaToolCrash(RuntimeFault):
    error_code: RuntimeErrorCode = "RPA_TOOL_CRASH"
class VerificationMismatch(RuntimeFault):
    error_code: RuntimeErrorCode = "VERIFICATION_MISMATCH"
class VerificationTimeout(RuntimeFault):
    error_code: RuntimeErrorCode = "VERIFICATION_TIMEOUT"
class PostHandoverCrash(RuntimeFault):
    error_code: RuntimeErrorCode = "POST_HANDOVER_CRASH"
    

# ============================================================
# BACKENDS
# ============================================================

try: from custom_backends import build_backends # type: ignore
except ImportError:
    build_backends = None


class DemoMailBackend:
    """Demo mailbox simulated with local folders and .eml files (replace w/ eg. Outlook)"""

    MAIL_STATUS_PREFIXES = ("PROCESSING", "DONE", "FAIL")

    def __init__(self, logger, source_type) -> None:
        self.logger = logger
        self.source_type = source_type # change to folder in e.g. outlook
        self.inbox_dir = Path(self.source_type) / "inbox"
        self.inbox_dir.mkdir(parents=True, exist_ok=True)

    def list_inbox_mail_paths(self, max_items=None) -> list[str]:
        paths_raw = sorted(self.inbox_dir.glob("*.eml"))

        if max_items is not None:
            paths_raw = paths_raw[:max_items]

        paths = [str(x) for x in paths_raw] #convert Path-type to str
        return paths

    def parse_mail_file(self, mail_path) -> JobCandidate:
        with open(mail_path, "rb") as f:
            msg = BytesParser(policy=policy.default).parse(f)

        from_name, from_address = parseaddr(msg.get("From", ""))
        del from_name # not used

        email_address = (from_address or "").strip().lower()
        if not email_address or "@" not in email_address:
            email_address = None

        email_subject = msg.get("Subject", "").strip()

        # not needed in demo
        # message_id = msg.get("Message-ID", "").strip()    # eg. Outlook EntryID / Graph ID in real backend
        # raw_headers = {k: str(v) for k, v in msg.items()} # good for troubleshooting metadata 

        if msg.is_multipart():
            body_parts = []
            for part in msg.walk():
                if part.get_content_type() == "text/plain" and not part.get_filename():
                    try:
                        body_parts.append(part.get_content())
                    except Exception:
                        pass
            email_body = "\n".join(body_parts).strip()
        else:
            try:
                email_body = msg.get_content().strip()
            except Exception:
                email_body = ""
        

        attachments = {}
        # placeholder for implementation
        #attachments = {
        #    "attachments": [
        #        {
        #            "filename": "orders.xlsx",
        #            "path": "/some/path/orders.xlsx",
        #        }
        #    ]
        #}
    
        return JobCandidate(
            source_ref=mail_path,
            email_address=email_address,
            email_subject=email_subject,
            email_body=email_body,
            source_type=self.source_type,
            parsed_source_data=attachments,
            )

    def mark_processing(self, candidate: JobCandidate, job_id: int | None = None) -> JobCandidate:
        original_subject = self._strip_status_prefix(candidate.email_subject)
        new_subject = f"PROCESSING/{self._today_yyyymmdd()}/{original_subject}"
        return self._set_subject(candidate, new_subject, job_id)

    def mark_done(self, candidate: JobCandidate, job_id: int | None = None) -> JobCandidate:
        original_subject = self._strip_status_prefix(candidate.email_subject)
        new_subject = f"DONE/{self._today_yyyymmdd()}/{original_subject}"
        return self._set_subject(candidate, new_subject, job_id)

    def mark_failed(self, candidate: JobCandidate, job_id: int | None = None) -> JobCandidate:
        original_subject = self._strip_status_prefix(candidate.email_subject)
        new_subject = f"FAIL/{self._today_yyyymmdd()}/{original_subject}"
        return self._set_subject(candidate, new_subject, job_id)

    def send_reply(self, candidate: JobCandidate, extra_subject: str, extra_body: str, job_id: int) -> None:

        reply_to = candidate.email_address

        original_subject = self._strip_status_prefix(candidate.email_subject)
        subject = f"{extra_subject} re: {original_subject}"

        body = (
            f"{extra_body} \n\n"
            f"-------------------------------------------------------------\n"
            f"{candidate.email_body}"
        ) # In a real mail backend, this should use the native reply mechanism.

        self.logger.system(f"message sent, with body={body[:100]}... (GDPR sanitized)", job_id)
        
        assert reply_to is not None
        self._print_email_preview(reply_to, subject, body)
        
    def delete(self, candidate: JobCandidate, job_id: int | None = None) -> None:
        self.logger.system(f"removing: {candidate.source_ref}", job_id)
        os.remove(candidate.source_ref)

    def _today_yyyymmdd(self) -> str:
        return datetime.datetime.now().strftime("%Y%m%d")

    def _strip_status_prefix(self, subject: str | None) -> str:
        subject = (subject or "").strip()

        for status in self.MAIL_STATUS_PREFIXES:
            pattern = rf"^{status}/\d{{8}}/(.*)$"
            match = re.match(pattern, subject, flags=re.IGNORECASE)
            if match:
                return match.group(1).strip()

        return subject

    def _has_status_prefix(self, candidate: JobCandidate, status: str | None = None) -> bool:
        subject = (candidate.email_subject or "").strip()

        if status is not None:
            return bool(re.match(rf"^{status}/\d{{8}}/", subject, flags=re.IGNORECASE))

        return bool(re.match(r"^(PROCESSING|DONE|FAIL)/\d{8}/", subject, flags=re.IGNORECASE))

    def _set_subject(self, candidate: JobCandidate, new_subject: str, job_id: int | None = None) -> JobCandidate:
        """ Demo backend: update Subject inside the .eml file, real Outlook backend: use native subject rename."""
        path = Path(candidate.source_ref)

        with open(path, "rb") as f:
            msg = BytesParser(policy=policy.default).parse(f)

        if "Subject" in msg:
            msg.replace_header("Subject", new_subject)
        else:
            msg["Subject"] = new_subject

        with open(path, "wb") as f:
            f.write(msg.as_bytes(policy=policy.default))

        candidate.email_subject = new_subject
        self.logger.system(f"renamed mail subject to {new_subject}", job_id)

        return candidate

    def _print_email_preview(self, reply_to: str, subject: str, body: str):

        print(
        "\n" + "="*72 +
        "\n📧 EMAIL REPLY PREVIEW\n" +
        "="*72 +
        f"\nFrom:    robot@runtime.local"
        f"\nTo:      {reply_to}"
        f"\nSubject: {subject}"
        f"\nDate:    {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        "\n" + "-"*72 +
        f"\n{body}\n" +
        "="*72 + "\n"
    )


class DemoErpBackend:
    """Demo ERP backend simulated with an Excel file."""

    def order_adjust_selection_rows(self, path="Demo_ERP_table.xlsx") -> list[dict]:
        # do a well targeted 'query' 

        self._ensure_demo_erp_exists(path)

        try:
            wb = load_workbook(path)
        except BadZipFile:
            time.sleep(1)
            wb = load_workbook(path)

        ws = wb.active

        assert ws is not None #to satisfy pylance

        all_rows=[]

        for row in ws.iter_rows(min_row=2):  # skip header
            
            source_ref = row[0].value
            order_qty = row[1].value
            material_available = row[2].value

            if order_qty != material_available:

                all_rows.append({
                        "source_ref": source_ref,
                        "order_qty": order_qty,
                        "material_available": material_available,
                    })
                
        wb.close()
        return all_rows
    
    def build_candidate_from_row(self, row) -> JobCandidate:
            
        source_ref = row.get("source_ref")
        order_qty = row.get("order_qty")
        material_available = row.get("material_available")


        try: order_qty = int(order_qty)
        except Exception: raise ValueError(f"invalid order_qty: {order_qty}")
        try: material_available = int(material_available)
        except Exception: raise ValueError(f"invalid material_available: {material_available}")


        parsed_source_data ={
            "order_qty": order_qty,
            "material_available": material_available,
        }

        return JobCandidate(
            source_ref=str(source_ref),
            source_type="erp_query",
            parsed_source_data=parsed_source_data
        )
        
    def get_order_qty(self, source_ref, path="Demo_ERP_table.xlsx") -> int | None:
        self._ensure_demo_erp_exists(path)

        try:
            wb = load_workbook(path)
        except BadZipFile:
            time.sleep(1)
            wb = load_workbook(path)

        ws = wb.active
        assert ws is not None #to satisfy pylance

        for row in ws.iter_rows(min_row=2):
            cell_source_ref = row[0].value

            if str(cell_source_ref) == str(source_ref):
                value = row[1].value  # order_qty

                if isinstance(value, int):
                    wb.close()
                    return int(value)
                
                else: 
                    raise ValueError(f"order_qty: {value} is not INT")
        
        wb.close()
        return None

    def _ensure_demo_erp_exists(self, path="Demo_ERP_table.xlsx") -> None:
        """Create the demo ERP table if it does not exist."""
        if os.path.exists(path):
            return

        wb = Workbook()
        ws = wb.active
        assert ws is not None #to satisfy pylance

        # headers
        ws["A1"] = "source_ref"
        ws["B1"] = "order_qty"
        ws["C1"] = "material_available"

        wb.save(path)
        wb.close()


# ============================================================
# JOB FLOWS
# ============================================================

class MailFlow:
    """Handle email-driven job intake."""
    def __init__(self, logger, friends_repo, audit, is_within_operating_hours, network_service, personal_mail_handlers, shared_mail_handlers, job_lifecycle, personal_mailbox, shared_mailbox) -> None:
        self.logger = logger
        self.friends_repo = friends_repo
        self.audit = audit
        self._is_within_operating_hours = is_within_operating_hours
        self.network_service = network_service
        self.personal_mail_handlers = personal_mail_handlers
        self.shared_mail_handlers = shared_mail_handlers
        self.job_lifecycle = job_lifecycle
        self.personal_mailbox = personal_mailbox
        self.shared_mailbox = shared_mailbox
        self.friends_filename = Path(self.friends_repo.friends_path).name

    def poll_once(self):
        if self._poll_personal_once():
            return True
        if self._poll_shared_once():
            return True
        return False

    def _poll_personal_once(self) -> bool:
        ''' direct human-to-robot channel (parse, always claim) '''

        candidate: JobCandidate

        personal_inbox_paths = self.personal_mailbox.list_inbox_mail_paths()
        if personal_inbox_paths:
            self.friends_repo.reload_if_modified()

        for path in personal_inbox_paths:
            candidate = self.personal_mailbox.parse_mail_file(path)
            
            if self.personal_mailbox._has_status_prefix(candidate, "PROCESSING"):
                raise PreHandoverCrash("stale personal mail found with PROCESSING subject prefix",)

            self._handle_personal_mail(candidate) 
            return True

        return False
    
    def _poll_shared_once(self) -> bool:
        """External business mailbox: parse, skip irrelevant, handle one in-scope mail."""

        if not self._is_within_operating_hours():
            return False

        if not self.network_service.has_network_access():
            return False

        shared_inbox_paths = self.shared_mailbox.list_inbox_mail_paths()

        for path in shared_inbox_paths:
            candidate = self.shared_mailbox.parse_mail_file(path)

            # stale processing mail = previous crash/interruption
            if self.shared_mailbox._has_status_prefix(candidate, "PROCESSING"):
                self.shared_mailbox.mark_failed(candidate)
                self.logger.system(f"stale shared mail marked FAIL: {candidate.source_ref}")
                self.logger.ui("--> returned (stale processing mail)")
                return True

            # already handled shared mail
            if self.shared_mailbox._has_status_prefix(candidate, "DONE"):
                continue

            if self.shared_mailbox._has_status_prefix(candidate, "FAIL"):
                continue

            # not our job
            handler = self._find_shared_mail_handler(candidate)
            if handler is None:
                continue
          
            self._handle_shared_mail(candidate, handler)
            return True

        return False

    def _handle_personal_mail(self, candidate: JobCandidate) -> None:
        if self.personal_mailbox._has_status_prefix(candidate, "PROCESSING"):
            raise PreHandoverCrash("stale personal mail found with PROCESSING subject prefix")

        self.logger.system(f"{candidate.source_type} produced mail {candidate.source_ref}")
        self.logger.ui(f"email from {candidate.email_address}", blank_line_before=True)

        if not self.friends_repo.is_allowed_sender(candidate.email_address):
            self.job_lifecycle.delete_only(
                candidate,
                ui_log=f"--> rejected (not in {self.friends_filename})",
                system_log=f"--> rejected (not in {self.friends_filename})",
            )
            return

        if not self._is_within_operating_hours():
            self.job_lifecycle.reject_personal_mail(
                candidate,
                error_code="OUTSIDE_WORKING_HOURS",
                reason="Outside robot's working hours 05-23.",
                ui_log="--> rejected (outside working hours)",
            )
            return

        handler = self._find_personal_mail_handler(candidate)
        if handler is None:
            self.job_lifecycle.reject_personal_mail(
                candidate,
                error_code="UNKNOWN_JOB",
                reason="Could not identify a job type from your email.",
                ui_log="--> rejected (unable to identify job type)",
            )
            return

        job_name = handler.job_name

        if not self.friends_repo.has_job_access(candidate.email_address, job_name):
            self.job_lifecycle.reject_personal_mail(
                candidate,
                error_code="NO_ACCESS",
                reason=f"Request denied, your email is not permitted to trigger '{job_name}'.",
                job_name=job_name,
                ui_log=f"--> rejected (no access to {job_name})",
            )
            return

        if not self.network_service.has_network_access():
            self.job_lifecycle.reject_personal_mail(
                candidate,
                error_code="NO_NETWORK",
                reason="No network connection at the moment.",
                job_name=job_name,
                ui_log="--> rejected (no network connection)",
            )
            return

        result = handler.precheck_and_build_payload(candidate)
        if not result.is_success:
            self.job_lifecycle.reject_personal_mail(
                candidate,
                error_code="INVALID_INPUT",
                reason=result.error_message,
                job_name=job_name,
                ui_log=f"--> rejected (invalid input for {job_name})",
            )
            return

        self.job_lifecycle.queue_for_rpa(
            candidate,
            job_name=job_name,
            rpatool_payload=result.rpatool_payload or {},
            send_online_notice=True,
            start_recording=True,
        )

    def _find_personal_mail_handler(self, mail: JobCandidate):
        for handler in self.personal_mail_handlers.values():
            if handler.can_handle(mail):
                return handler
        return None

    def _handle_shared_mail(self, candidate: JobCandidate, handler) -> None:
        self.logger.system("shared inbox mail detected (sender/subject omitted)")
        self.logger.ui(f"email (shared) from {candidate.email_address}", blank_line_before=True,)


        job_name = handler.job_name

        try:
            result = handler.precheck_and_build_payload(candidate)

        except Exception as err:
            self.job_lifecycle.skip_shared_mail(
                candidate,
                error_code="PRE_HANDOVER_CRASH",
                reason=f"Unhandled error in shared mail handler: {err}",
                job_name=job_name,
                ui_log=f"--> returned (code error in {job_name})",
            )
            return

        if not result.is_success:
            self.job_lifecycle.skip_shared_mail(
                candidate,
                error_code=result.error_code or "INVALID_INPUT",
                reason=result.error_message or "Invalid input.",
                job_name=job_name,
                ui_log=f"--> returned (invalid input for {job_name})",
            )
            return

        self.job_lifecycle.queue_for_rpa(
            candidate,
            job_name=job_name,
            rpatool_payload=result.rpatool_payload or {},
            send_online_notice=False,
            start_recording=True,
        )

    def _find_shared_mail_handler(self, mail: JobCandidate):
        for handler in self.shared_mail_handlers.values():
            if handler.can_handle(mail):
                return handler
        return None


class QueryFlow:
    """Handle query-driven job intake."""

    def __init__(self, logger, query_handlers, job_lifecycle, is_within_operating_hours) -> None:
        self.logger = logger
        self.query_handlers = query_handlers
        self.job_lifecycle = job_lifecycle
        self._is_within_operating_hours = is_within_operating_hours

    def poll_once(self) -> bool:

        work_item: QueryWorkItem

        if not self._is_within_operating_hours():
            return False

        for handler in self.query_handlers.values():
            #self.logger.system(f"checking query handler {handler.job_name}")
            work_item = handler.find_next_work_item()
            if work_item is None:
                continue

            self.logger.ui(
                f"query job detected: {work_item.candidate.source_ref}",
                blank_line_before=True,
            )

            self.job_lifecycle.queue_for_rpa(
                candidate=work_item.candidate,
                job_name=handler.job_name,
                rpatool_payload=work_item.rpatool_payload,
                send_online_notice=False,
                start_recording=True,
                # ui_log=f"--> accepted ({handler.job_name})",
            )
            return True

        return False


# ============================================================
# JOB HANDLERS
# ============================================================

class PingHandler:
    '''This 'automation' allows the user to check if the robot (RobotRuntime + RPA tool) running.'''

    job_name: JobName = "ping"

    def __init__(self, logger) -> None:
        self.logger = logger

    def can_handle(self, candidate: JobCandidate) -> bool:
            subject = str(candidate.email_subject).strip()
            original_subject = re.sub(r"^(PROCESSING|DONE|FAIL)/\d{8}/", "", subject, flags=re.IGNORECASE).strip().lower()
            return original_subject == self.job_name

    def precheck_and_build_payload(self, candidate: JobCandidate) -> JobResult:
        return JobResult(is_success=True, rpatool_payload={})

    def verify_result(self, handover_file: HandoverFile) -> JobResult:
        return JobResult(is_success=True)

try: from custom_query_jobs import build_custom_query_handlers
except ImportError: build_custom_query_handlers = None

try: from custom_personal_mail_jobs import build_custom_personal_mail_handlers
except ImportError: build_custom_personal_mail_handlers = None

try: from custom_shared_mail_jobs import build_custom_shared_mail_handlers
except ImportError: build_custom_shared_mail_handlers = None


# ============================================================
# HANDOVER
# ============================================================

class HandoverRepository:
    """Persist and validate the file-based state shared with the RPA tool."""

    def __init__(self, logger, handover_file) -> None:
        self.logger = logger
        self.handover_file = handover_file


    def read(self) -> HandoverFile:
        ''' read HANDOVER_FILE '''
        
        last_err=None

        for attempt in range(7):
            try:
                # read file
                with open(self.handover_file, "r", encoding="utf-8") as f:
                    handover_data = json.load(f)
                
                # rebuild object
                handover_file = self._validate_and_build_handover_file(handover_data)

                return handover_file
                
            except Exception as err:
                last_err = err
                self.logger.system(f"WARN: retry {attempt+1}/7 : {err}")
                time.sleep(attempt/10)
        
        
        raise RuntimeError(f"{self.handover_file} unreadable: {last_err}")
    
      
    def write(self, handover_file: HandoverFile) -> None:
        ''' atomic write of HANDOVER_FILE '''

        handover_data = asdict(handover_file)

        self._validate_and_build_handover_file(handover_data) # only validate (ignore return)
        job_id = handover_data.get("job_id")

        last_err = None
        
        for attempt in range(7):
            temp_path = None
            try:
                
                dir_path = os.path.dirname(os.path.abspath(self.handover_file))
                fd, temp_path = tempfile.mkstemp(dir=dir_path, suffix=".tmp")

                #atomic write
                with os.fdopen(fd, "w", encoding="utf-8") as tmp:
                    json.dump(handover_data, tmp, indent=2) # indent for human eyes
                    tmp.flush()
                    os.fsync(tmp.fileno())

                os.replace(temp_path, self.handover_file)
                
                self.logger.system(
                    f"wrote handover state={handover_data.get('state')}, job_name={handover_data.get('job_name')}, "
                    f"rpatool_payload={handover_data.get('rpatool_payload')} etc. (GDPR sanitized)",
                    job_id,
                )               
                return

            except Exception as err:
                last_err = err
                self.logger.system(f"WARN: {attempt+1}/7 error", job_id)
                time.sleep(attempt/10) # 0 0.1... 0.6 sec     

            finally:
                if temp_path and os.path.exists(temp_path):
                    try: os.remove(temp_path)
                    except Exception: pass

        self.logger.system(f"CRITICAL: cannot write {self.handover_file} {last_err}", job_id)
        raise RuntimeError(f"CRITICAL: cannot write {self.handover_file}")


    def _validate_and_build_handover_file(self, handover_data: dict) -> HandoverFile:
        """Validate raw handover dict and return HandoverFile."""

        state = handover_data.get("state")
        job_id = handover_data.get("job_id")
        job_name = handover_data.get("job_name")
        source_type = handover_data.get("source_type")
        source_ref = handover_data.get("source_ref")
        email_address = handover_data.get("email_address")
        email_subject = handover_data.get("email_subject")
        email_body = handover_data.get("email_body")
        parsed_source_data = handover_data.get("parsed_source_data")
        rpatool_payload = handover_data.get("rpatool_payload")

        if state is None:
            raise ValueError("state missing")

        if state not in get_args(HandoverState):
            raise ValueError(f"unknown state: {state}")

        if job_id is not None:
            try:
                job_id = int(job_id)
            except Exception:
                raise ValueError(f"job_id not INT-like: {job_id}")

        if state == "idle":
            if any(v is not None for v in (
                job_id, job_name, source_type, source_ref,
                email_address, email_subject, email_body, parsed_source_data, rpatool_payload
            )):
                raise ValueError(f"state 'idle' should have no more variables: {handover_data}")

        elif state in ("job_queued", "job_running", "job_verifying"):
            required_fields = {
                "job_id": job_id,
                "job_name": job_name,
                "source_type": source_type,
                "source_ref": source_ref,
                "parsed_source_data": parsed_source_data,
                "rpatool_payload": rpatool_payload,
            }

            missing = [k for k, v in required_fields.items() if v is None]
            if missing:
                raise ValueError(f"{state} has missing fields in {self.handover_file}: {missing}")

            if source_type not in get_args(SourceType):
                raise ValueError(f"unknown source_type: {source_type}")

            if source_type in ("personal_inbox", "shared_inbox"):
                required_fields = {
                    "email_address": email_address,
                    "email_subject": email_subject,
                    "email_body": email_body,
                }

                missing = [k for k, v in required_fields.items() if v is None]
                if missing:
                    raise ValueError(f"{source_type} has missing fields in {self.handover_file}: {missing}")
                
            if not isinstance(parsed_source_data, dict):
                raise ValueError("parsed_source_data must be dict")
            if not isinstance(rpatool_payload, dict):
                raise ValueError("rpatool_payload must be dict")
        
        elif state == "safestop":
            pass

        return HandoverFile(
            state=state,
            job_id=job_id,
            job_name=job_name,
            source_type=source_type,
            source_ref=source_ref,
            email_address=email_address,
            email_subject=email_subject,
            email_body=email_body,
            parsed_source_data=parsed_source_data,
            rpatool_payload=rpatool_payload,
        )


    def is_valid_observed_transition(self, prev_state: HandoverState | None, state: HandoverState) -> bool:
        """Validate transitions observed by polling. Allows skipped states."""

        if prev_state is None: # at startup
            return True

        allowed_observed: dict[HandoverState, set[HandoverState]] = {
            "idle": {"job_queued", "job_running", "job_verifying", "safestop"},
            "job_queued": {"job_running", "job_verifying", "safestop"},
            "job_running": {"job_verifying", "safestop"},
            "job_verifying": {"idle", "safestop"},
            "safestop": {"idle"},
        }

        allowed_next = allowed_observed[prev_state]

        return state in allowed_next
       

class JobLifecycleService:
    """Execute pre-handover actions and build HandoverFile objects for the RPA tool."""

    def __init__(self, logger, handover, show_recording_overlay, recording, audit, notifications, personal_mailbox, shared_mailbox, job_handlers, hide_recording_overlay, generate_job_id) -> None:
        self.logger = logger
        self.handover = handover
        self.recording = recording
        self.audit = audit
        self._show_recording_overlay = show_recording_overlay
        self.notifications = notifications
        self.personal_mailbox = personal_mailbox
        self.shared_mailbox = shared_mailbox
        self.job_handlers = job_handlers
        self._hide_recording_overlay = hide_recording_overlay
        self.generate_job_id = generate_job_id


    def delete_only(self, candidate: JobCandidate, ui_log: str, system_log: str | None = None) -> None:
        self.logger.ui(ui_log)
        self.logger.system(system_log)
        self.personal_mailbox.delete(candidate)

    def reject_personal_mail(self, candidate: JobCandidate, error_code: RuntimeErrorCode, reason: str | None, job_name: JobName | None = None, ui_log: str | None = None,) -> None:
        if ui_log:
            self.logger.ui(ui_log)

        job_id = self.generate_job_id()

        self.audit.insert(
            job_id=job_id,
            source_ref=candidate.source_ref,
            email_address=candidate.email_address,
            email_subject=candidate.email_subject,
            job_name=job_name,
            started_at_date=datetime.datetime.now().strftime("%Y-%m-%d"),
            started_at_time=datetime.datetime.now().strftime("%H:%M:%S"),
            lifecycle_status="REJECTED",
            source_type=candidate.source_type,
            error_code=error_code,
            error_message=reason,
        )

        self.notifications.send_final_reply_and_delete_original(
            candidate=candidate,
            lifecycle_status="REJECTED",
            error_code=error_code,
            job_id=job_id,
            reason=reason,
        )

        self.audit.update(job_id=job_id, final_reply_sent=True)

    def skip_shared_mail(self, candidate: JobCandidate, error_code: RuntimeErrorCode, reason: str, job_name: JobName | None = None, ui_log: str | None = None,) -> None:
        if ui_log:
            self.logger.ui(ui_log)

        job_id = self.generate_job_id()
        
        self.audit.insert(
            job_id=job_id,
            source_ref=candidate.source_ref,
            email_address=candidate.email_address,
            email_subject=candidate.email_subject,
            job_name=job_name,
            started_at_date=datetime.datetime.now().strftime("%Y-%m-%d"),
            started_at_time=datetime.datetime.now().strftime("%H:%M:%S"),
            lifecycle_status="REJECTED",
            source_type=candidate.source_type,
            error_code=error_code,
            error_message=reason,
        )

        self.shared_mailbox.mark_failed(candidate, job_id)

    def queue_for_rpa(self, candidate: JobCandidate, job_name: JobName, rpatool_payload: dict[str, Any], send_online_notice: bool, start_recording: bool, ui_log: str | None = None,) -> None:
        if ui_log:
            self.logger.ui(ui_log)

        job_id = self.generate_job_id()

        if candidate.source_type == "personal_inbox":
            candidate = self.personal_mailbox.mark_processing(candidate, job_id)

        elif candidate.source_type == "shared_inbox":
            candidate = self.shared_mailbox.mark_processing(candidate, job_id)

        self.audit.insert(
            job_id=job_id,
            source_ref=candidate.source_ref,
            email_address=candidate.email_address,
            email_subject=candidate.email_subject,
            job_name=job_name,
            started_at_date=datetime.datetime.now().strftime("%Y-%m-%d"),
            started_at_time=datetime.datetime.now().strftime("%H:%M:%S"),
            lifecycle_status="QUEUED",
            source_type=candidate.source_type,
        )

        if start_recording:
            self._maybe_start_recording(job_id)

        if send_online_notice:
            if not self.audit.has_sender_job_today(candidate.email_address, job_id):
                self.notifications.send_online_notice(candidate, job_id)

        handover = HandoverFile(
            state="job_queued",
            job_id=job_id,
            job_name=job_name,
            source_type=candidate.source_type,
            source_ref=candidate.source_ref,
            email_address=candidate.email_address,
            email_subject=candidate.email_subject,
            email_body=candidate.email_body,
            parsed_source_data=candidate.parsed_source_data,
            rpatool_payload=rpatool_payload,
        )

        self.handover.write(handover)

    def complete_from_handover(self, handover_file: HandoverFile) -> None:
        job_id = handover_file.job_id

        self.audit.mark_verifying(job_id)
        result = self.validate(handover_file)
        
        self._hide_recording_overlay()
        self.recording.stop(job_id)
        self.recording.try_upload_recording(job_id)

        if result.is_success:
            self.audit.mark_done(job_id)
            self._finalize(handover_file, "DONE")
        else:
            self.audit.mark_failed(job_id, result.error_code, result.error_message)
            self._finalize(handover_file, "FAIL", result.error_code, result.error_message)

    def _maybe_start_recording(self, job_id: int|None):
        
        started = self.recording.start(job_id)
        if started:
            try: self._show_recording_overlay()
            except Exception as e: self.logger.system(f"error {e}", job_id)

    def _validate_verification_result(self, verification_result: JobResult):
        if verification_result.is_success:
            if verification_result.error_code:
                raise ValueError(f"error_code must be empty for is_success=True, is {verification_result.error_code}")
            if verification_result.error_message:
                raise ValueError(f"error_message must be empty for is_success=True, is {verification_result.error_message}")
            return
            
        allowed_error_codes = {"VERIFICATION_MISMATCH", "VERIFICATION_TIMEOUT"}
        if verification_result.error_code not in allowed_error_codes:
            raise ValueError(f"unknown verification_result.error_code={verification_result.error_code}. Allowed: {allowed_error_codes}")
        if not verification_result.error_message:
            raise ValueError("missing verification_result.error_message for is_success=False")

    def validate(self, handover_file: HandoverFile) -> JobResult:
        job_id = handover_file.job_id
        job_name = handover_file.job_name
        
        self.logger.system(f"completing {job_name} with payload {handover_file.rpatool_payload}", job_id) # only store safe data in log

        try:            
            handler = self.job_handlers.get(job_name)
            if handler is None:
                raise PostHandoverCrash(
                    f"No handler for job_name={job_name}",
                    job_id=job_id,
                    handover_file=handover_file,
                )
            
            verification_result = handler.verify_result(handover_file)
            self._validate_verification_result(verification_result)

            return verification_result

            
        except RuntimeFault:
            raise
            
        except Exception as err:
            try:
                self.audit.update(
                    job_id=job_id,
                    lifecycle_status="FAIL",
                    error_code="POST_HANDOVER_CRASH",
                    error_message=f"crash during verification stage: {err}",
            )
            except Exception as err2:
                self.logger.system(f"[PostHandoverService] {err} {err2}", job_id)
            

            raise PostHandoverCrash(
                f"verification stage crashed, outcome unknown: {err}",
                job_id=job_id,
                handover_file=handover_file,
                cause=err,
            ) from err

    def _update_logs(self, lifecycle_status: str, handover_file: HandoverFile,) -> None:
        job_name = handover_file.job_name

        self.logger.ui(f"--> {lifecycle_status.lower()} ({job_name})")
        self.logger.system(f"{lifecycle_status} ({job_name})", handover_file.job_id)
   
    def _finalize(self, handover_file: HandoverFile, lifecycle_status: LifecycleStatus, error_code: str | None=None, error_message: str | None=None) -> None:
        job_id = handover_file.job_id

      
        if error_code == "VERIFICATION_TIMEOUT":
            assert error_message is not None # to satisfy pylance
            raise VerificationTimeout(
                error_message,
                job_id=job_id,
                handover_file=handover_file,
            )

        if error_code == "VERIFICATION_MISMATCH":
            assert error_message is not None # to satisfy pylance
            raise VerificationMismatch(
                error_message,
                job_id=job_id,
                handover_file=handover_file,
            )
 
        # do mail specifics
        final_reply_sent = self._handle_mail_completion(handover_file, lifecycle_status, error_code, error_message)

        if final_reply_sent:
            self.audit.update(job_id=job_id, final_reply_sent=True,)

        self._update_logs(lifecycle_status, handover_file)
  
    def _build_candidate_from_handover(self, handover_file: HandoverFile) -> JobCandidate:
        assert handover_file.source_ref is not None # to satisfy pylance
        assert handover_file.parsed_source_data is not None # to satisfy pylance
        assert handover_file.source_type is not None # to satisfy pylance
        
        return JobCandidate(
                source_ref=handover_file.source_ref,
                source_type=handover_file.source_type,
                parsed_source_data=handover_file.parsed_source_data,
                email_address=handover_file.email_address,
                email_subject=handover_file.email_subject,
                email_body=handover_file.email_body,
                )
        
    def _handle_mail_completion(self, handover_file: HandoverFile, lifecycle_status: LifecycleStatus, error_code: str | None, error_message: str | None) -> bool:

        if handover_file.source_type not in ("personal_inbox", "shared_inbox"):
            return False
    
        # rebuild candidate
        candidate = self._build_candidate_from_handover(handover_file)

        job_id = handover_file.job_id

        if handover_file.source_type == "personal_inbox":
            if lifecycle_status == "DONE":
                self.notifications.send_final_reply_and_delete_original(
                    candidate=candidate,
                    lifecycle_status=lifecycle_status,
                    job_id=job_id,
                    error_code=None,
                )
                return True

            if lifecycle_status == "FAIL":
                if error_code in {"VERIFICATION_MISMATCH", "VERIFICATION_TIMEOUT"}:
                    # Safestop recovery sends the final user reply with out-of-service context.
                    return False
                
                self.notifications.send_final_reply_and_delete_original(
                    candidate=candidate,
                    lifecycle_status=lifecycle_status,
                    error_code=error_code,
                    job_id=job_id,
                    reason=error_message,
                )
                return True

            raise ValueError(f"unexpected personal inbox lifecycle_status={lifecycle_status}")


        if handover_file.source_type == "shared_inbox":
            if lifecycle_status == "DONE":
                self.shared_mailbox.mark_done(candidate, job_id)
                return False

            if lifecycle_status == "FAIL":
                self.shared_mailbox.mark_failed(candidate, job_id)
                return False

            raise ValueError(f"unexpected shared inbox lifecycle_status={lifecycle_status}")
       
        return False


# ============================================================
# USER NOTIFICATIONS
# ============================================================

class UserNotificationService:
    """Only for personal_inbox user-facing replies."""

    def __init__(self, personal_mailbox, recordings_destination_folder, rpa_tool_claim_timeout, rpa_tool_execution_timeout, rpa_admin_email):
        self.personal_mailbox = personal_mailbox
        self.recordings_destination_folder = recordings_destination_folder
        self.rpa_tool_claim_timeout = rpa_tool_claim_timeout
        self.rpa_tool_execution_timeout = rpa_tool_execution_timeout
        self.rpa_admin_email = rpa_admin_email
        
        self.MAILCOMMAND_JOB_ID = 999999999999


    def send_final_reply_and_delete_original(self, candidate: JobCandidate, lifecycle_status: LifecycleStatus, error_code: str | None, job_id: int, reply_context = "runtime", reason=None, delete_after=True) -> None:
        
        subject, body = self._build_job_reply(
            lifecycle_status=lifecycle_status,
            error_code=error_code,
            job_id=job_id,
            reason=reason,
            reply_context=reply_context,
        )

        self._send(candidate, subject, body, job_id, delete_after)


    def send_out_of_service_reply(self, candidate: JobCandidate, job_id: int) -> None:
        
        self.send_final_reply_and_delete_original(
            candidate=candidate,
            lifecycle_status="FAIL",
            error_code="OUT_OF_SERVICE",
            job_id=job_id,
            )


    def send_command_reply(self, candidate: JobCandidate) -> None:
        self._send(
            candidate=candidate,
            subject="got it!",
            body="Command received.",
            job_id=self.MAILCOMMAND_JOB_ID,
            delete_after=True,
        )


    def send_admin_alert(self, reason: str) -> None:
        fake_candidate = JobCandidate(
            source_ref="safestop, no real source_ref",
            email_address=self.rpa_admin_email,
            email_subject="",
            email_body="",
            source_type="personal_inbox",
            parsed_source_data={},
        )

        body = (
            "Robot is in degraded mode.\n\n"
            f"Reason:\n{reason}\n\n"
            "Reminder of available email commands: 'stop1234' and 'restart1234'."
        )

        self._send(
            candidate=fake_candidate,
            subject="safestop notice",
            body=body,
            job_id=self.MAILCOMMAND_JOB_ID,
            delete_after=False,
        )


    def send_online_notice(self, candidate: JobCandidate, job_id: int) -> None:
        # TODO add average completion time calculated from job_audit for requested job_name 

        body = (
            ">Hello, human<\n\n"
            "The first request each day is replied with: online\n"
            "You should receive a final reply after completion\n"
            f"(in max {self.rpa_tool_claim_timeout + self.rpa_tool_execution_timeout} seconds from now)."
        )
       
        body += self._get_robot_signature()

        self._send(
            candidate=candidate,
            subject="ONLINE",
            body=body,
            job_id=job_id,
            delete_after=False,
        )


    def _classify_reply_kind(self, lifecycle_status:LifecycleStatus, error_code: str | None) -> str:

        if lifecycle_status == "DONE":
            return "DONE"
        
        if lifecycle_status == "FAIL" and error_code == "PRE_HANDOVER_CRASH":
            return "NOT_STARTED"

        if lifecycle_status == "FAIL" and error_code == "OUT_OF_SERVICE":
            return "OUT_OF_SERVICE"

        if lifecycle_status == "FAIL" and error_code == "RPA_TOOL_CRASH":
            return "STARTED_BUT_CRASHED"

        if lifecycle_status == "FAIL" and error_code == "VERIFICATION_MISMATCH":
            return "VERIFICATION_MISMATCH"
        
        if lifecycle_status == "FAIL" and error_code == "VERIFICATION_TIMEOUT":
            return "VERIFYING_CRASH"

        if lifecycle_status == "FAIL" and error_code == "POST_HANDOVER_CRASH":
            return "VERIFYING_CRASH"

        if lifecycle_status == "REJECTED":
            return "NOT_STARTED"

        if lifecycle_status == "QUEUED":
            return "NOT_STARTED"

        if lifecycle_status == "RUNNING":
            return "STARTED_BUT_CRASHED"

        if lifecycle_status == "VERIFYING":
            return "VERIFYING_CRASH"

        if lifecycle_status == "FAIL":
            return "UNKNOWN_FAIL"

        raise ValueError(f"Cannot classify reply for lifecycle_status={lifecycle_status}, error_code={error_code}")


    def _build_job_reply(self, lifecycle_status: LifecycleStatus, error_code: str | None, job_id: int, reason, reply_context: str,) -> tuple[str, str]:
        # TODO: for increased user value, extend reply with a short summary eg. "changed PO 450221 on SKU 110212 from 34pcs to 31pcs"

        subject: str
        body: str

        recording_text = self._get_recording_text(job_id)
        reply_kind = self._classify_reply_kind(lifecycle_status, error_code)

        if reply_kind == "DONE":
            subject = "DONE"
            body = (
                    f"Job completed successfully.\n\n"
                    f"job_id: {job_id}\n\n"
                    f"{recording_text}"
                    f"This email can be deleted."
                )

        elif reply_kind == "NOT_STARTED":
            subject = "FAIL"
            body = (
                    f"Your request was not started.\n\n"
                    f"{f'Reason: {reason}\n\n' if reason else ''}"
                    #f"job_id: {job_id}\n"
                    f"Keep calm, no changes were made in ERP.\n"
                    f"This email can be deleted."
                )

        elif reply_kind == "STARTED_BUT_CRASHED":
            subject = "FAIL"
            body = (
                    f"The robot started your request, but then crashed.\n\n"
                    f"{f'Reason: {reason}\n\n' if reason else ''}"
                    f"job_id: {job_id}\n"
                    f"Changes may have been made in ERP before the crash.\n"
                    f"Please (very recommended) review the result manually.\n\n"
                    f"{recording_text}"
                    f"This email can be deleted."
                )
                
        elif reply_kind == "VERIFICATION_MISMATCH":
            subject = "FAIL"
            body = (
                    f"The robot completed the request, and the result was checked in ERP.\n"
                    f"However, the final ERP data did not match the expected result.\n\n"
                    f"{f'Reason: {reason}\n\n' if reason else ''}"
                    f"job_id: {job_id}\n"
                    f"Please (very recommended) review the result manually.\n\n"
                    f"{recording_text}"
                    f"This email can be deleted."
                )

        elif reply_kind == "VERIFYING_CRASH":
            subject = "FAIL"
            body = (
                    f"The robot completed the request, but crashed during the final verification stage.\n"
                    f"The outcome could therefore not be confirmed automatically.\n\n"
                    f"job_id: {job_id}\n"
                    f"Please verify the result manually in ERP.\n\n"
                    f"{recording_text}"
                    f"This email can be deleted."
                )

        elif reply_kind == "OUT_OF_SERVICE":
            subject = "FAIL"
            body = (
                    "The robot is temporary out-of-service and does not accept any new requests.\n"
                    #"This email can be deleted."
                )

        elif reply_kind == "UNKNOWN_FAIL":
            subject = "FAIL"
            body = (
                    f"The robot crashed and the exact job outcome could not be classified.\n\n"
                    f"{f'Reason: {reason}\n\n' if reason else ''}"
                    f"job_id: {job_id}\n"
                    f"Please review the result manually in ERP.\n\n"
                    f"{recording_text}"
                    f"This email can be deleted."
                )

        else:
            raise ValueError(f"Unhandled reply_kind={reply_kind}")
        

        if reply_context == "safestop_recovery":
            body = body.replace(
                    "This email can be deleted.",

                    "To avoid further problems, the robot will temporary go out-of-service and also notify robot admin to fix the issue.\n"
                    "This email can be deleted.",
                )

            if reply_kind == "VERIFICATION_MISMATCH":
                # avoid confusion that admin will “correct the mismatch” 
                body = body.replace(
                    " to fix the issue.\n",
                    ".\n"
                )

        elif reply_context == "startup_recovery":
            body = (
                    "The robot was offline and has now restarted.\n"
                    "If you already received a final reply (DONE/FAIL) for this job, you can ignore this recovery message."
                ) + "\n\n" + body
        
        body += self._get_robot_signature()

        return subject, body

   
    def _get_recording_text(self, job_id: int) -> str:
        recording_path = Path(self.recordings_destination_folder) / f"{job_id}.mp4"
        if recording_path.exists():
            return (
                "A screen recording is available for review:\n"
                f"{recording_path}\n\n"
            )

        return ""
    

    def _get_robot_signature(self) -> str:
        return (
            "\n\n---\n"
            "Automated message from Robot.\n"
        )


    def _send(self, candidate: JobCandidate, subject: str, body: str, job_id: int, delete_after: bool,) -> None:
        self.personal_mailbox.send_reply(
            candidate=candidate,
            extra_subject=subject,
            extra_body=body,
            job_id=job_id,
        )

        if delete_after:
            self.personal_mailbox.delete(
                candidate=candidate,
                job_id=job_id,
            )


# ============================================================
# RECORDING / SAFESTOP / INFRASTRUCTURE
# ============================================================   
                      
class RecordingService:
    ''' screen-recorder to capture all RPA tool screen-activity '''

    def __init__(self, logger, recordings_in_progress_folder, recordings_destination_folder) -> None:
        self.logger = logger
        self.recording_process = None
        self._ffmpeg_warned = False
        self.recordings_in_progress_folder = recordings_in_progress_folder
        self.recordings_destination_folder = recordings_destination_folder


    def _get_screen_resolution(self):
        # written by AI chatgpt 5.3 instant
        try:
            output = subprocess.check_output(["xrandr"], text=True)
            for line in output.splitlines():
                if "*" in line:
                    res = line.split()[0]
                    return res.split("x")
        except Exception:
            pass

        # fallback: Tkinter
        try:
            root = tk.Tk()
            root.withdraw()
            width = root.winfo_screenwidth()
            height = root.winfo_screenheight()
            root.destroy()
            return str(width), str(height)
        except Exception:
            pass

        return "1920", "1080"

 
    def start(self, job_id) -> bool:
        """start the screen recording"""
        # written by AI chatgpt 5.3 instant
        try:

            os.makedirs(self.recordings_in_progress_folder, exist_ok=True)
            filename = f"{self.recordings_in_progress_folder}/{job_id}.mp4"

            drawtext = (
                f"drawtext=text='job_id  {job_id}':"
                "x=200:y=20:"
                "fontsize=32:"
                "fontcolor=lightyellow:"
                "box=1:"
                "boxcolor=black@0.5"
            )

            if platform.system() == "Windows":
                ffmpeg_path = None

                local_ffmpeg = Path("./ffmpeg.exe")
                if local_ffmpeg.exists():
                    ffmpeg_path = str(local_ffmpeg)
                else:
                    ffmpeg_in_path = shutil.which("ffmpeg")
                    if ffmpeg_in_path:
                        ffmpeg_path = ffmpeg_in_path

                if ffmpeg_path is None:
                    if not self._ffmpeg_warned:
                        message = (
                            "FFMPEG NOT FOUND\n\n"
                            "Screen recording is disabled.\n\n"
                            "Fix:\n"
                            "1. Go to: https://www.gyan.dev/ffmpeg/builds/\n"
                            "2. Download 'ffmpeg-git-essentials'\n"
                            "3. Extract the archive\n"
                            "4. Open the 'bin' folder\n"
                            "5. Copy ffmpeg.exe next to main.py\n"
                        )

                        print("\n" + "="*60 + "\n" + message + "\n" + "="*60 + "\n")
                        self.logger.system(message, job_id)
                        self.logger.ui("--> recording disabled (ffmpeg missing)")
                        self._ffmpeg_warned = True
                    return False

                capture = ["-f", "gdigrab", "-i", "desktop"]

                recording_process = subprocess.Popen(
                    [
                        ffmpeg_path,
                        "-y",
                        *capture,
                        "-framerate", "15",
                        "-vf", drawtext,
                        "-vcodec", "libx264",
                        "-pix_fmt", "yuv420p",
                        "-preset", "ultrafast",
                        filename,
                    ],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    creationflags=getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0),
                )

            else:
                display = os.environ.get("DISPLAY")
                if not display:
                    self.logger.system("WARN: screen-recording disabled because DISPLAY is missing", job_id)
                    return False

                ffmpeg_path = shutil.which("ffmpeg")
                if ffmpeg_path is None:
                    self.logger.system("WARN: screen-recording disabled because ffmpeg is not installed", job_id)
                    return False

                width, height = self._get_screen_resolution()

                capture = [
                    "-video_size", f"{width}x{height}",
                    "-f", "x11grab",
                    "-i", display,
                ]

                recording_process = subprocess.Popen(
                    [
                        ffmpeg_path,
                        "-y",
                        *capture,
                        "-framerate", "15",
                        "-vf", drawtext,
                        "-vcodec", "libx264",
                        "-preset", "ultrafast",
                        filename,
                    ],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    start_new_session=True,
                )

            time.sleep(0.2)
            
        except Exception as e:
            # treat a screen recording crash, when screen recording should be used, as a critical error
            raise PreHandoverCrash("Unable to start screen recording", job_id=job_id, cause=e) from e

        if recording_process.poll() is not None:
            self.logger.system("WARN: ffmpeg exited immediately; recording did not start", job_id)
            raise PreHandoverCrash("Unable to start screen recording", job_id=job_id, cause=None)
        
        self.recording_process = recording_process
        self.logger.system("recording started", job_id)
        return True

        
    def stop(self, job_id=None) -> None:
        ''' allow global kill of FFMPEG processes since Orchestrator is designed to run on a dedicated machine '''
        # written by AI chatgpt 5.3 instant

        try: self.logger.system("stop recording", job_id)
        except Exception: pass

        recording_process = self.recording_process
        self.recording_process = None

        try:
            if recording_process is not None:
                # try first stop only our own process
                if platform.system() == "Windows":
                    try:
                        recording_process.send_signal(
                            getattr(signal, "CTRL_BREAK_EVENT", signal.SIGTERM)
                        )
                    except Exception:
                        try:
                            recording_process.terminate()
                        except Exception:
                            pass

                    try:
                        recording_process.wait(timeout=8)
                        return
                    except subprocess.TimeoutExpired:
                        pass

                    # else, kill only our own process
                    try:
                        subprocess.run(
                            ["taskkill", "/PID", str(recording_process.pid), "/T", "/F"],
                            stdout=subprocess.DEVNULL,
                            stderr=subprocess.DEVNULL,
                            check=False,
                        )
                        recording_process.wait(timeout=3)
                        return
                    except Exception:
                        pass

                    # last resort, global kill all ffmpeg
                    subprocess.run(
                        ["taskkill", "/IM", "ffmpeg.exe", "/T", "/F"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        check=False,
                    )

                else:
                    # try first stop only our own process
                    try:
                        os.killpg(recording_process.pid, signal.SIGINT)

                    except Exception:
                        try:
                            recording_process.terminate()
                        except Exception:
                            pass

                    try:
                        recording_process.wait(timeout=8)
                        return
                    except subprocess.TimeoutExpired:
                        pass

                    # else, kill only our own process
                    try:
                        os.killpg(recording_process.pid, signal.SIGKILL)
                        recording_process.wait(timeout=3)
                        return
                    except Exception:
                        pass

                    # last resort, global kill all ffmpeg
                    subprocess.run(
                        ["killall", "-q", "-KILL", "ffmpeg"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        check=False,
                    )

            else:
                # fallback if process object is lost
                if platform.system() == "Windows":
                    subprocess.run(
                        ["taskkill", "/IM", "ffmpeg.exe", "/T", "/F"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        check=False,
                    )
                else:
                    subprocess.run(
                        ["killall", "-q", "-KILL", "ffmpeg"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        check=False,
                    )
        except Exception as err:
            self.logger.system(f"WARN from stop(): {err}", job_id)


    def try_upload_recording(self, job_id, max_attempts=3) -> None:
        ''' upload to a shared drive'''
    
        local_file = f"{self.recordings_in_progress_folder}/{job_id}.mp4"
        local_file = Path(local_file)

        if not local_file.exists():
            self.logger.system(f"no recording file found to upload", job_id)
            return
        
        remote_path = Path(self.recordings_destination_folder) / f"{job_id}.mp4"
        remote_path.parent.mkdir(parents=True, exist_ok=True)

        for attempt in range(max_attempts):
            try:
                
                shutil.copy2(local_file, remote_path)
                self.logger.system(f"upload successful: {remote_path}", job_id)
                try: os.remove(local_file)
                except Exception: pass

                return

            except Exception as e:
                self.logger.system(f"upload attempt {attempt+1}/{max_attempts} failed: {e}", job_id)
                time.sleep(attempt + 1)
        
        self.logger.system(f"upload failed: {remote_path}", job_id)


    def cleanup_aborted_recordings(self):
        """Upload or clean up recordings left behind by aborted runs."""

        directory = Path(self.recordings_in_progress_folder)
        if not directory.exists():
            return
        
        for file in directory.iterdir():

            if file.is_file() and file.suffix == ".mp4":
                job_id = file.stem
                self.logger.system(f"cleanup upload started", job_id)
                self.try_upload_recording(job_id)


class FriendsRepository:
    '''Access-control source for personal_inbox'''

    def __init__(self, friends_path, allowed_job_names) -> None:
        self.friends_path = friends_path
        self.friends_filename = Path(friends_path).name
        self.allowed_job_names = allowed_job_names
        self.access_by_email: dict[str, set[str]] = {}
        self.access_file_mtime: float | None = None

    def _ensure_friends_file_exists(self) -> None:
        '''Create a template access file if missing.'''
        if os.path.exists(self.friends_path):
            return

        wb = Workbook()
        ws = wb.active
        assert ws is not None

        ws["A1"] = "email"
        ws["B1"] = "ping"
        ws["C1"] = "qty_adjust"

        ws["A2"] = "alice@example.com"
        ws["B2"] = "x"

        ws["A3"] = "bob@test.com"
        ws["B3"] = "x"
        ws["C3"] = "x"

        wb.save(self.friends_path)
        wb.close()

    def _load_access_file(self) -> dict[str, set[str]]:
        '''
        Reads access file and returns eg:

        {
            "alice@example.com": {"ping"},
            "bob@test.com": {"ping", "qty_adjust"}
        }
        '''
        # written by AI chatgpt 5.3 instant

        wb = load_workbook(self.friends_path, data_only=True)
        try:
            ws = wb.active
            assert ws is not None

            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                raise ValueError(f"{self.friends_filename} contains no users")

            header = rows[0]
            self._validate_friends_header(header)
            access_map: dict[str, set[str]] = {}

            for row in rows[1:]:
                email_cell = row[0]
                if email_cell is None:
                    continue

                email = str(email_cell).strip().lower()
                if not email:
                    continue

                permissions: set[str] = set()

                for col in range(1, len(header)):
                    jobname = header[col]
                    if jobname is None:
                        continue

                    jobname = str(jobname).strip().lower()
                    cell = row[col] if col < len(row) else None

                    if cell is None:
                        continue

                    if str(cell).strip().lower() == "x":
                        permissions.add(jobname)

                access_map[email] = permissions

            return access_map
        finally:
            wb.close()

    def reload_if_modified(self) -> bool:
        '''Reload access file if changed.'''
        # written by AI chatgpt 5.3 instant

        self._ensure_friends_file_exists()

        mtime = os.path.getmtime(self.friends_path)
        if self.access_file_mtime == mtime:
            return False

        new_access = self._load_access_file()
        self._validate_friends_access(new_access)

        self.access_by_email = new_access
        self.access_file_mtime = mtime

        return True

    def is_allowed_sender(self, email_address: str | None) -> bool:

        if not email_address:
            return False
        
        email = email_address.strip().lower()        
        return email in self.access_by_email

    def has_job_access(self, email_address: str, job_name: str) -> bool:
        email = email_address.strip().lower()
        job = job_name.strip().lower()
        return job in self.access_by_email.get(email, set())

    def _validate_friends_access(self, access_map: dict[str, set[str]]) -> None:
        if not isinstance(access_map, dict):
            raise ValueError("access_map must be dict")

        valid_job_names = self.allowed_job_names

        for email, permissions in access_map.items():
            if not isinstance(email, str):
                raise ValueError(f"invalid email key type: {email}")

            email_normalized = email.strip().lower()
            if not email_normalized:
                raise ValueError("empty email in access_map")

            if "@" not in email_normalized:
                raise ValueError(f"invalid email in {self.friends_filename}: {email}")

            if not isinstance(permissions, set):
                raise ValueError(f"permissions must be set for {email}")

            invalid_permissions = permissions - valid_job_names
            if invalid_permissions:
                print(
                    f"WARN! invalid job types for {email}: {sorted(invalid_permissions)}. "
                    f"Allowed: {sorted(valid_job_names)}"
                )
            
    def _validate_friends_header(self, header_row) -> None:
        if not header_row or str(header_row[0]).strip().lower() != "email":
            raise ValueError(f"{self.friends_filename} column A must be 'email'")

        valid_job_names = self.allowed_job_names

        for col in range(1, len(header_row)):
            jobname = header_row[col]
            if jobname is None:
                continue

            jobname_str = str(jobname).strip().lower()
            if jobname_str not in valid_job_names:
                print(
                    f"WARN! invalid job type column in {self.friends_filename}: {jobname_str}. "
                    f"Allowed: {sorted(valid_job_names)}"
                )


class NetworkService:
    """Check whether the machine currently has access to the required company network resources."""


    def __init__(self, logger, network_healthcheck_path) -> None:
        self.logger = logger
        self.network_state = False
        self.next_network_check_time = 0
        self.network_healthcheck_path = network_healthcheck_path


    def has_network_access(self) -> bool:

        now = time.time()

        if now < self.next_network_check_time:
            return self.network_state

        try:
            if self.network_healthcheck_path is None: # demo assumption
                online = True                         # demo assumption
            else:
                os.listdir(self.network_healthcheck_path)
                online = True 

        except Exception:
            online = False
            
        if online != self.network_state:
            self.network_state = online

            if online:
                self.logger.system("network restored")
            else:
                self.logger.system(f"WARN: network lost")

        # check once every minute if offline, once every hour if online
        if online:
            self.next_network_check_time = now + 3600
        else:
            self.next_network_check_time = now + 60
        
        return online


class AuditRepository:
    ''' handles an audit-style activity log '''

    def __init__(self, logger, audit_db_path) -> None:
        self.logger = logger
        self.audit_db_path = audit_db_path

    def _connect_with_retry(self) -> sqlite3.Connection:
  
        max_retries = 3
        for attempt in range(max_retries):
            try:
                conn = sqlite3.connect(self.audit_db_path, timeout=10)
                return conn
            except sqlite3.OperationalError as e:
                self.logger.system(f"WARN: {e}")
                if attempt == max_retries - 1:
                    raise
                time.sleep(0.5)

        # below will never execute?
        conn = sqlite3.connect(self.audit_db_path, timeout=10,)
        return conn  
        
    def ensure_db_exists(self) -> None:
        
        with self._connect_with_retry() as conn:
            cur = conn.cursor()
           
            cur.execute('''
                CREATE TABLE IF NOT EXISTS audit_log
                         (
                        job_id INTEGER PRIMARY KEY, 
                        job_name TEXT, 
                        lifecycle_status TEXT, 
                        email_address TEXT, 
                        email_subject TEXT, 
                        source_ref TEXT,
                        started_at_date TEXT, 
                        started_at_time TEXT, 
                        updated_at_time TEXT, 
                        final_reply_sent INTEGER NOT NULL DEFAULT 0,
                        source_type TEXT,
                        error_code TEXT, 
                        error_message TEXT 
                        )
                        ''')

    def _build_audit_fields(self, job_id, email_address=None, email_subject=None, source_ref=None, job_name: JobName | None = None, started_at_date=None, started_at_time=None, lifecycle_status: LifecycleStatus | None = None, final_reply_sent=None, source_type: SourceType | None = None, error_code=None, error_message=None,) -> dict:
        all_fields = {
            "job_id": job_id,
            "email_address": email_address,
            "email_subject": email_subject,
            "source_ref": source_ref,
            "job_name": job_name,
            "started_at_date": started_at_date,
            "started_at_time": started_at_time,
            "updated_at_time": datetime.datetime.now().strftime("%H:%M:%S"),
            "lifecycle_status": lifecycle_status,
            "final_reply_sent": final_reply_sent,
            "source_type": source_type,
            "error_code": error_code,
            "error_message": error_message,
        }

        # drop None:s
        fields = {k: v for k, v in all_fields.items() if v is not None}

        gdpr_safe_fields = dict(fields)
        
        if gdpr_safe_fields.get("email_address") is not None: gdpr_safe_fields["email_address"] = "***"
        if gdpr_safe_fields.get("email_subject") is not None: gdpr_safe_fields["email_subject"] = "***"

        suffix = " (GDPR-sanitized)" if fields != gdpr_safe_fields else ""
        self.logger.system(f"received audit fields {gdpr_safe_fields}{suffix}", job_id)

        return fields

    def insert(self, job_id, email_address=None, email_subject=None, source_ref=None, job_name: JobName | None=None, started_at_date=None, started_at_time=None, lifecycle_status: LifecycleStatus | None=None, final_reply_sent=None, source_type:SourceType | None=None, error_code=None, error_message=None,) -> None:
        # use for new row

        fields = self._build_audit_fields(
            job_id=job_id,
            email_address=email_address,
            email_subject=email_subject,
            source_ref=source_ref,
            job_name=job_name,
            started_at_date=started_at_date,
            started_at_time=started_at_time,
            lifecycle_status=lifecycle_status,
            final_reply_sent=final_reply_sent,
            source_type=source_type,
            error_code=error_code,
            error_message=error_message,
        )
        
        columns = ", ".join(fields.keys())
        placeholders = ", ".join("?" for _ in fields)

        with self._connect_with_retry() as conn:
            cur = conn.cursor()

            cur.execute(
                f"INSERT INTO audit_log ({columns}) VALUES ({placeholders})",
                tuple(fields.values())
            )

    def update(self, job_id, email_address=None, email_subject=None, source_ref=None, job_name: JobName | None=None, started_at_date=None, started_at_time=None, lifecycle_status: LifecycleStatus | None=None, final_reply_sent=None, source_type:SourceType | None=None, error_code=None, error_message=None,) -> None:
        # use eg: self.audit.update(job_id=20260311124501, job_name="qty_adjust")

        fields = self._build_audit_fields(
            job_id=job_id,
            email_address=email_address,
            email_subject=email_subject,
            source_ref=source_ref,
            job_name=job_name,
            started_at_date=started_at_date,
            started_at_time=started_at_time,
            lifecycle_status=lifecycle_status,
            final_reply_sent=final_reply_sent,
            source_type=source_type,
            error_code=error_code,
            error_message=error_message,
        )
        
        fields.pop("job_id", None)

        if not fields:
            return

        set_clause = ", ".join(f"{k}=?" for k in fields)

        with self._connect_with_retry() as conn:
            cur = conn.cursor()

            cur.execute(
                f"UPDATE audit_log SET {set_clause} WHERE job_id=?",
                (*fields.values(), job_id)
            )

            if cur.rowcount == 0:
                raise ValueError(f"update(): no row in DB with job_id={job_id}")

    def count_done_jobs_today(self) -> int:
        today = datetime.date.today().isoformat()

        with self._connect_with_retry() as conn:
            cur = conn.cursor()
            cur.execute('''
                SELECT COUNT(*)
                FROM audit_log
                WHERE started_at_date = ?
                AND lifecycle_status = 'DONE'
            ''', (today,))
            
            result = cur.fetchone()[0]

        return result

    def has_sender_job_today(self, sender_mail, job_id) -> bool:

        today = datetime.datetime.now().strftime("%Y-%m-%d")
        with self._connect_with_retry() as conn:
            cur = conn.cursor()

            cur.execute(
                '''
                SELECT COUNT(*)
                FROM audit_log
                WHERE started_at_date = ? AND email_address = ? AND job_id != ?
                ''',
                (today, sender_mail, job_id,)
            )

            jobs_today = cur.fetchone()[0]

        return jobs_today > 0

    def has_been_processed_today(self, source_ref) -> bool:
        # use to avoid bad loops in query-jobs

        today = datetime.datetime.now().strftime("%Y-%m-%d")
        with self._connect_with_retry() as conn:
            cur = conn.cursor()

            cur.execute(
                '''
                SELECT COUNT(*)
                FROM audit_log
                WHERE started_at_date = ? AND source_ref = ?
                ''',
                (today, source_ref,)
            )

            jobs_today = cur.fetchone()[0]

        return jobs_today > 0

    def get_latest_job_id(self) -> int:
        with self._connect_with_retry() as conn:
            cur = conn.cursor()
            cur.execute('''
                SELECT job_id
                FROM audit_log
                ORDER BY job_id DESC
                LIMIT 1
            ''')
            row = cur.fetchone()

        return row[0] if row is not None else 0

    def get_personal_pending_reply_jobs(self) -> list[dict]:
        source_type: SourceType = "personal_inbox"

        with self._connect_with_retry() as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute(
                '''
                SELECT job_id, source_type, email_address, email_subject, source_ref, lifecycle_status, error_code, error_message
                FROM audit_log
                WHERE source_type = ?
                AND COALESCE(final_reply_sent, 0) = 0
                ORDER BY job_id
                ''',
                (source_type,)
            )
            rows = cur.fetchall()

        list_of_dicts = [dict(row) for row in rows]

        return list_of_dicts

    def get_job_by_source_ref(self, source_ref: str) -> dict | None:
        with self._connect_with_retry() as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute(
                '''
                SELECT *
                FROM audit_log
                WHERE source_ref = ?
                ORDER BY job_id DESC
                LIMIT 1
                ''',
                (source_ref,)
            )
            row = cur.fetchone()

        return dict(row) if row is not None else None

    def mark_verifying(self, job_id):
        self.update(
            job_id=job_id, 
            lifecycle_status="VERIFYING",
            )
        
    def mark_done(self, job_id):
        self.update(
            job_id=job_id, 
            lifecycle_status="DONE", 
            )
    
    def mark_failed(self, job_id, error_code, error_message):
        self.update(
            job_id=job_id, 
            lifecycle_status="FAIL", 
            error_code=error_code, 
            error_message=error_message, 
            )

class LoggerService:
    """ logging functions """

    def __init__(self, dashboard_ui, system_log_path) -> None:
        self.dashboard_ui = dashboard_ui
        self.system_log_path = system_log_path

    def ui(self, text:str, blank_line_before: bool = False) -> None:
        
        self.dashboard_ui.post_log_line(text, blank_line_before)

    def system(self, event_text, job_id: int | None=None,):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        event_text = str(event_text)

        # get caller function name
        try:
            frame = sys._getframe(1)
            caller_name = frame.f_code.co_name
            instance = frame.f_locals.get("self")
            if instance is not None:
                class_name = instance.__class__.__name__
                caller = f"{class_name}.{caller_name}()"
            else:
                caller = f"{caller_name}()"

        except Exception:
            caller = "unknown_caller()"
      
        log_line = f"{timestamp} | py  | job_id={job_id or ''} | {caller} | {event_text}"

        # normalize to single-line
        log_line = " ".join(str(log_line).split())

        last_err = None
        for i in range(7):
            try:
                with open(self.system_log_path, "a", encoding="utf-8") as f:
                    f.write(log_line + "\n")
                    f.flush()
                return

            except Exception as err:
                last_err = err
                print(f"WARN: retry {i+1}/7 from log_system():", err)
                time.sleep(i + 1)

        # fallback to print() when log fails        
        print(f"[print fallback] {job_id} {event_text} | {last_err}")  
 

class MailRecoveryService:
    '''stuck/crashed mail handler used in while in safestop'''


    def __init__(self, logger, personal_mailbox, shared_mailbox, audit, friends_repo, notifications, generate_job_id) -> None:
        self.logger = logger
        self.personal_mailbox = personal_mailbox
        self.shared_mailbox = shared_mailbox
        self.audit = audit
        self.job_audit_filename = Path(self.audit.audit_db_path).name
        self.friends_repo = friends_repo
        self.notifications = notifications
        self.friends_filename = Path(self.friends_repo.friends_path).name
        self.generate_job_id = generate_job_id
   
    def process_one_personal_mail_in_safestop(self) -> None:
        """Process one personal inbox email while in safestop."""

        # process one personal inbox email in degraded mode
        paths = self.personal_mailbox.list_inbox_mail_paths(max_items=1)
        if not paths:
            return
        
        inbox_path = paths[0]
        candidate = self.personal_mailbox.parse_mail_file(inbox_path)                

        try: self.logger.ui(f"email from {candidate.email_address}", blank_line_before=True)
        except Exception: pass

        # silent delete non friends
        if not self.friends_repo.is_allowed_sender(candidate.email_address):
            try: self.logger.ui(f"--> rejected (not in {self.friends_filename})")
            except Exception: pass
            self.personal_mailbox.delete(candidate)
            return
        
        # check for email commands
        if self._check_for_stop_command(candidate):
            return
        if self._check_for_restart_command(candidate):
            return

        # reply, audit-log and delete for friends
        job_id = self.generate_job_id()
        final_reply_sent = self._try_send_out_of_service_reply(candidate, job_id)
        try: self.insert_recovery_audit_row(job_id, candidate, final_reply_sent, recovery_reason="SAFESTOP")
        except Exception as e: self.logger.system(f"audit error={e}")
        
        try: self.logger.ui("--> rejected (safestop)")
        except Exception: pass

        return
  
    def _try_send_out_of_service_reply(self, candidate: JobCandidate, job_id: int) -> bool:
        final_reply_sent = False

        try:
            self.notifications.send_out_of_service_reply(candidate, job_id)
            final_reply_sent = True

        except Exception as e:
            self.logger.system(e, job_id)
            
        return final_reply_sent

    def _mark_faulted_pending_job_for_recovery(self, fault: RuntimeFault):
        # update audit row to FAIL for pending reply jobs
        job_id = fault.job_id
        error_code = fault.error_code
        error_message = fault.error_message

        if job_id is not None:
            for audit_row in self.audit.get_personal_pending_reply_jobs():
                if job_id == audit_row.get("job_id"):
                    try: self.audit.update(job_id=job_id, lifecycle_status="FAIL", error_code=error_code, error_message=error_message)
                    except Exception as e: self.logger.system(e, job_id)
    
    def recover_pending_personal_mail_replies(self, reply_context: str, fault: RuntimeFault | None = None,) -> None:
        # written by AI chatgpt 5.3 instant

        if reply_context not in ("safestop_recovery", "startup_recovery"):
            raise ValueError(f"unknown reply_context={reply_context}")

        if fault is not None:
            self._mark_faulted_pending_job_for_recovery(fault)

        fallback_items: list[tuple[int, JobCandidate]] = []

        try:
            pending_jobs_before_scan = self.audit.get_personal_pending_reply_jobs()
            pending_source_refs = {
                row["source_ref"]
                for row in pending_jobs_before_scan
                if row.get("source_ref") is not None
            }
            audit_available = True

        except Exception as e:
            self.logger.system(f"WARN: audit unavailable during recovery scan: {e}")
            pending_source_refs = set()
            audit_available = False

        paths = self.personal_mailbox.list_inbox_mail_paths()
        
        for processing_path in paths:
            try:
                candidate = self.personal_mailbox.parse_mail_file(processing_path)
            except Exception as e:
                self.logger.system(f"could not parse processing mail {processing_path}: {e}")
                continue

            if not self.personal_mailbox._has_status_prefix(candidate, "PROCESSING"):
                continue

            if not self.friends_repo.is_allowed_sender(candidate.email_address):
                try:
                    self.logger.ui(f"email recovered from {candidate.email_address}", blank_line_before=True)
                    self.logger.ui(f"--> rejected (not in {self.friends_filename})")
                except Exception:
                    pass

                try:
                    self.personal_mailbox.delete(candidate)
                except Exception as e:
                    self.logger.system(f"could not delete non-friend processing mail {candidate.source_ref}: {e}")
                continue

            if audit_available and candidate.source_ref in pending_source_refs:
                continue

            job_id = self.generate_job_id()

            if not audit_available:
                fallback_items.append((job_id, candidate))
                continue

            try:
                self.insert_recovery_audit_row(
                    job_id,
                    candidate,
                    final_reply_sent=False,
                    recovery_reason="RECOVERY",
                )
            except Exception as e:
                self.logger.system(f"WARN: audit insert failed during recovery: {e}", job_id)
                fallback_items.append((job_id, candidate))
                continue

            try:
                self.logger.ui(f"email recovered from {candidate.email_address}", blank_line_before=True)
                self.logger.ui(f"--> rejected ({reply_context} recovery)")
            except Exception:
                pass

            self.logger.system(f"email recovered from {candidate.email_address}", job_id)

        # No audit fallback for mails that could not be registered.
        for job_id, candidate in fallback_items:
            audit_row_artificial = {
                "lifecycle_status": "FAIL",
                "error_code": "PRE_HANDOVER_CRASH",
                "job_id": job_id,
                "error_message": f"{self.job_audit_filename} is unavailable",
            }

            try:
                self.notifications.send_final_reply_and_delete_original(
                    candidate=candidate,
                    lifecycle_status=audit_row_artificial["lifecycle_status"],
                    error_code=audit_row_artificial.get("error_code"),
                    job_id=audit_row_artificial["job_id"],
                    reason=audit_row_artificial.get("error_message"),
                    reply_context=reply_context,
                    delete_after=True,
                )

                self.logger.system("recovery reply sent without audit", job_id)

                try:
                    self.logger.ui(f"email recovered from {candidate.email_address}", blank_line_before=True)
                    self.logger.ui(f"--> rejected ({self.job_audit_filename} unavailable)")
                except Exception:
                    pass

            except Exception as err:
                self.logger.system(f"recovery reply failed without audit: {err}", job_id)

        if not audit_available:
            return

        pending_jobs_after_scan = self.audit.get_personal_pending_reply_jobs()

        for audit_row in pending_jobs_after_scan:
            job_id = audit_row.get("job_id")
            source_ref = audit_row.get("source_ref")

            path = Path(source_ref)
            if path.exists():
                candidate = self.personal_mailbox.parse_mail_file(str(path))
                self.logger.system(
                    f"re-built candidate={candidate.source_ref, candidate.email_address, candidate.email_subject}",
                    job_id,
                )
                delete_after = True
            else:
                audit_row["error_code"] = "RECOVERY_SOURCE_MISSING"
                self.audit.update(job_id=job_id, error_code="RECOVERY_SOURCE_MISSING")
                candidate = self._build_candidate_from_audit(audit_row)
                self.logger.system(
                    f"re-built candidate={candidate} from audit due to missing processing file {source_ref}",
                    job_id,
                )
                delete_after = False

            try:
                self.notifications.send_final_reply_and_delete_original(
                    candidate=candidate,
                    lifecycle_status=audit_row["lifecycle_status"],
                    error_code=audit_row.get("error_code"),
                    job_id=audit_row["job_id"],
                    reason=audit_row.get("error_message"),
                    reply_context=reply_context,
                    delete_after=delete_after,
                )
                self.logger.system("recovery reply sent", job_id)

                self.audit.update(
                    job_id=job_id,
                    final_reply_sent=True,
                )

            except Exception as err:
                self.logger.system(f"recovery reply failed: {err}", job_id)
                
    def recover_stuck_shared_mail(self, fault: RuntimeFault | None = None) -> None:
        # written by AI
        if fault is None:
            return

        handover_file = fault.handover_file

        if (
            handover_file is None
            or handover_file.source_type != "shared_inbox"
            or handover_file.source_ref is None
        ):
            return

        job_id = fault.job_id

        candidate = JobCandidate(
            source_ref=handover_file.source_ref,
            source_type="shared_inbox",
            parsed_source_data=handover_file.parsed_source_data or {},
            email_address=handover_file.email_address,
            email_subject=handover_file.email_subject,
            email_body=handover_file.email_body,
        )

        if job_id is not None:
            try:
                self.audit.update(
                    job_id=job_id,
                    lifecycle_status="FAIL",
                    error_code=fault.error_code,
                    error_message=fault.error_message,
                )
            except Exception as e:
                self.logger.system(f"shared audit update failed for faulted job: {e}", job_id)

        try:
            self.shared_mailbox.mark_failed(candidate, job_id or self.generate_job_id())
            self.logger.ui("--> returned (shared mail marked FAIL)")
        except Exception as e:
            self.logger.system(f"shared mail recovery failed, error={e}", job_id)

    def insert_recovery_audit_row(self, job_id:int, candidate:JobCandidate, final_reply_sent: bool, recovery_reason,):
        
        if recovery_reason == "SAFESTOP":
            lifecycle_status="REJECTED"
            error_code="IN_SAFESTOP"
            error_message="not accepting new jobs in safestop"
        
        elif recovery_reason == "RECOVERY":
            lifecycle_status="FAIL"
            error_code="PRE_HANDOVER_CRASH"
            error_message="unknown, mail stuck with PROCESSING subject prefix"
        
        else:
            raise ValueError(f"unknown reason: {recovery_reason}")


        now = datetime.datetime.now()
        source_type: SourceType = "personal_inbox" 
        
        self.audit.insert(
            job_id=job_id,
            source_ref=candidate.source_ref,
            email_address=candidate.email_address,
            email_subject=candidate.email_subject,
            started_at_date=now.strftime("%Y-%m-%d"),
            started_at_time=now.strftime("%H:%M:%S"),
            lifecycle_status=lifecycle_status,
            error_code=error_code,
            error_message=error_message,
            source_type = source_type,
            final_reply_sent = final_reply_sent,
        )

    def _build_candidate_from_audit(self, audit_row) -> JobCandidate:
        
        return JobCandidate(
            source_ref = audit_row.get("source_ref"),
            source_type = audit_row.get("source_type"),
            email_address = audit_row.get("email_address"),
            email_subject = audit_row.get("email_subject"),
            email_body = "[ORIGINAL MESSAGE LOST]",
            parsed_source_data = {},
            )
  
    def _check_for_stop_command(self, candidate: JobCandidate) -> bool:

        if "stop1234" in str(candidate.email_subject).strip().lower():
            self.logger.system(f"stop command received from {candidate.email_address}")
            Path("stop.flag").write_text("", encoding="utf-8")
            
            try: self.notifications.send_command_reply(candidate)
            except Exception: pass
            return True
        
        return False

    def _check_for_restart_command(self, candidate: JobCandidate) -> bool:

        if "restart1234" in str(candidate.email_subject).strip().lower():
            self.logger.system(f"restart command received from {candidate.email_address}")
            Path("restart.flag").write_text("", encoding="utf-8")
            
            try: self.notifications.send_command_reply(candidate)
            except Exception: pass
            return True
        
        return False


class SafestopController:
    """Handle degraded mode, crash recovery, and operator restart/stop commands."""

    def __init__(self, logger, recording, hide_recording_overlay, post_status_update, set_ui_shutdown, check_for_stop_flag, handover_file, mail_recovery, notifications) -> None:
        self.logger = logger
        self.recording = recording
        self._hide_recording_overlay = hide_recording_overlay
        self._post_status_update = post_status_update
        self._set_ui_shutdown = set_ui_shutdown
        self._check_for_stop_flag = check_for_stop_flag
        self.HANDOVER_FILE = handover_file
        self.mail_recovery = mail_recovery
        self.notifications = notifications
        self._degraded_mode_entered = False

    def run_degraded_mode(self, fault: RuntimeFault,) -> None:
        '''
        Rules:
        * no job intake
        * mail-flow inactivated
        * query-flow inactivated
        * 'safestop' status text in UI
        * STOP and RESTART commands available 
        * REJECTED reply to new emails from users in friends.xlsx
        * changes in friends.xlsx access list will have no effect
        * notification email is sent to admin
        '''
        
        if self._degraded_mode_entered: return
        self._degraded_mode_entered = True

        job_id = fault.job_id
        handover_file = fault.handover_file

        crash_report = f"ROBOT RUNTIME CRASHED\n\nfault={fault}\n\n{fault.traceback_text}"
        
        if handover_file is not None and handover_file.state != "idle":
            crash_report += (
                f"\n\n...while working on job_name={handover_file.job_name} "
                f"with rpatool_payload=\n{handover_file.rpatool_payload}"
            )

            # overwrite only job_queued to stop RPA_tool from possibly start
            if handover_file.state == "job_queued":
                try:
                    handover_file.state="safestop"
                    self._write_handover_directly(handover_file)

                except Exception:
                    try: os.remove(self.HANDOVER_FILE)
                    except Exception as e: self.logger.system(e)
           
        self.logger.system(crash_report, job_id)

        try: self.recording.stop()
        except Exception as e: self.logger.system(e, job_id)

        try: self.recording.cleanup_aborted_recordings()
        except Exception as e: self.logger.system(e, job_id)

        try: self.notifications.send_admin_alert(crash_report)
        except Exception as e: self.logger.system(e, job_id)

        try:
            self.logger.ui(f"--> CRASH! All automations are stopped. Admin is notified")
        except Exception as e: self.logger.system(e, job_id)

        try: self.mail_recovery.recover_pending_personal_mail_replies(reply_context="safestop_recovery", fault=fault)
        except Exception as e: self.logger.system(e, job_id)

        try: self.mail_recovery.recover_stuck_shared_mail(fault)
        except Exception as e: self.logger.system(e, job_id)

        # placeholder for recovery logic for post_handover crash/mismatch for query jobs

        try: self._hide_recording_overlay()
        except Exception as e: self.logger.system(e, job_id)

        try: self._post_status_update("safestop")
        except Exception as e:
            self.logger.system(e, job_id)
            try: self._set_ui_shutdown()
            except Exception as e2: 
                self.logger.system(e2, job_id)
                os._exit(1)
            time.sleep(3)
            os._exit(0)
        
        self._enter_degraded_loop()
    
    def _check_for_restart_flag(self,) -> None:
        restartflag = "restart.flag"

        if os.path.isfile(restartflag):
            try: os.remove(restartflag)
            except Exception: pass
            self.logger.system(f"restart-command received from {restartflag}")
            
            try:
                self._write_handover_directly(HandoverFile(state="idle"))
            except Exception as e:
                self.logger.system(f"could not reset handover before restart: {e}")
                os._exit(1)
            
            self._restart_application()

    def _write_handover_directly(self, handover_file: HandoverFile) -> None:
        '''write w/o using full handover in degraded mode'''
        handover_data = asdict(handover_file)

        temp_path = f"{self.HANDOVER_FILE}.tmp"

        with open(temp_path, "w", encoding="utf-8") as f:
            json.dump(handover_data, f, indent=2)
            f.flush()
            os.fsync(f.fileno())

        os.replace(temp_path, self.HANDOVER_FILE)

    def _enter_degraded_loop(self) -> Never:
        ''' follow policy to always reply to known users'''  

        self.logger.system("running")

        while True:
            try:
                time.sleep(1)
                self._check_for_stop_flag()
                self._check_for_restart_flag()

                self.mail_recovery.process_one_personal_mail_in_safestop()


            except Exception as e:
                self.logger.system(e)

    def _restart_application(self) -> Never:
        ''' written by AI chatgpt 5.3 instant '''
        self.logger.system("restarting application in new visible terminal")

        try:
            self._set_ui_shutdown()
        except Exception:
            pass

        try:
            script_path = os.path.abspath(sys.argv[0])

            if platform.system() == "Windows":
                subprocess.Popen(
                    [sys.executable, script_path],
                    creationflags=subprocess.CREATE_NEW_CONSOLE # type: ignore
                )

            else:
                python_cmd = f'"{sys.executable}" "{script_path}"'

                terminal_candidates = [
                    ["gnome-terminal", "--", "bash", "-lc", f"{python_cmd}; exec bash"],
                    ["xfce4-terminal", "--hold", "-e", python_cmd],
                    ["konsole", "-e", "bash", "-lc", f"{python_cmd}; exec bash"],
                    ["xterm", "-hold", "-e", python_cmd],
                ]

                launched = False
                for cmd in terminal_candidates:
                    try:
                        subprocess.Popen(cmd)
                        launched = True
                        break
                    except FileNotFoundError:
                        continue

                if not launched:
                    raise RuntimeError("No supported terminal emulator found for restart")

        except Exception as e:
            self.logger.system(e)
            os._exit(1)

        time.sleep(1)
        os._exit(0)


# ============================================================
# UI
# ============================================================

class DashboardUI:
    """Tkinter dashboard for runtime status, logs, and operator visibility."""

    # colors
    BG = "#000000"
    TEXT = "#F5F5F5"
    MUTED = "#A0A0A0"
    GREEN = "#22C55E"
    GREEN_2 = "#16A34A"
    GREEN_3 = "#15803D"
    RED = "#DC2626"
    YELLOW = "#FACC15"
    SCROLL_TROUGH = "#0F172A"
    SCROLL_BG = "#1E293B"
    SCROLL_ACTIVE = "#475569"

    # fonts
    FONT_STATUS = ("Arial", 100, "bold")
    FONT_COUNTER = ("Segoe UI", 140, "bold")
    FONT_SMALL = ("Arial", 14, "bold")
    FONT_LOG = ("DejaVu Sans Mono", 20)
    FONT_RECORDING = ("Arial", 20, "bold")

    # sizes
    ROOT_PADX = 50
    SCROLLBAR_WIDTH = 23

    RECORDING_WIDTH = 250
    RECORDING_HEIGHT = 110
    RECORDING_MARGIN_RIGHT = 30


    def __init__(self):
        self._build_root(self.BG)
        self._build_header(self.BG, self.TEXT)
        self._build_body(self.BG, self.TEXT)
        self._build_footer(self.BG, self.TEXT)

        #self._debug_grid(self.root)

    def run(self) -> None:
        self.root.mainloop()

    def shutdown(self) -> None:
        if self._closing:
            return

        self._closing = True

        self.root.destroy()

    def _debug_grid(self, widget):
        ''' highlights all grids with red '''
        for child in widget.winfo_children():
            try: child.configure(highlightbackground="red", highlightthickness=1)
            except Exception: pass
            self._debug_grid(child)

    def _build_root(self, bg_color):
        self.root = tk.Tk()

        w = self.root.winfo_screenwidth()
        h = self.root.winfo_screenheight()
        self.root.geometry(f"{w}x{h}+0+0")

        self.root.configure(bg=bg_color, padx=self.ROOT_PADX)
        self._closing = False
        self.root.protocol("WM_DELETE_WINDOW", self._on_close_attempt)

        self.root.title('RPA dashboard')
        self._create_recording_overlay()

        # layout using grid
        self.root.grid_rowconfigure(1, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

    def _build_header(self, bg_color, text_color):
        self.header = tk.Frame(self.root, bg=bg_color)

        self.header.grid(row=0, column=0, sticky="ew")
        self.header.grid_columnconfigure(2, weight=1)
        self.header.grid_rowconfigure(0, weight=1)

        # Header content
        self.rpa_text_label = tk.Label(
            self.header,
            text="RPA:",
            fg=text_color,
            bg=bg_color,
            font=self.FONT_STATUS,
        )  
        self.rpa_text_label.grid(row=0, column=0, padx=16, pady=16, sticky="w")

        self.rpa_status_label = tk.Label(
            self.header,
            text="",
            fg=self.RED,
            bg=bg_color,
            font=self.FONT_STATUS,
        )
        self.rpa_status_label.grid(row=0, column=1, padx=16, pady=16, sticky="w")

        self.status_dot = tk.Label(
            self.header,
            text="",
            fg=self.GREEN,
            bg=bg_color,
            font=("Arial", 50, "bold"),
        )
        self.status_dot.grid(row=0, column=2, sticky="w")

        # jobs done today (counter + label in same grid)
        self.jobs_counter_frame = tk.Frame(self.header, bg=bg_color)
        self.jobs_counter_frame.grid(row=0, column=3, sticky="ne", padx=40, pady=30)
        self.jobs_counter_frame.grid_rowconfigure(0, weight=1)
        self.jobs_counter_frame.grid_columnconfigure(0, weight=1)

        # normal view (jobs done today)
        self.jobs_normal_view = tk.Frame(self.jobs_counter_frame, bg=bg_color)
        self.jobs_normal_view.grid(row=0, column=0, sticky="nsew")
        self.jobs_normal_view.grid_columnconfigure(0, weight=1)

        self.jobs_done_label = tk.Label(
            self.jobs_normal_view,
            text="0",
            fg=text_color,
            bg=bg_color,
            font=self.FONT_COUNTER,
            anchor="e",
            justify="right",
        )
        self.jobs_done_label.grid(row=0, column=0, sticky="e")

        self.jobs_counter_text = tk.Label(
            self.jobs_normal_view,
            text="jobs done today",
            fg=self.MUTED,
            bg=bg_color,
            font=self.FONT_SMALL,
            anchor="e",
        )
        self.jobs_counter_text.grid(row=1, column=0, sticky="e", pady=(0, 6))

        # safestop view (big X)
        self.jobs_error_view = tk.Frame(self.jobs_counter_frame, bg=bg_color)
        self.jobs_error_view.grid(row=0, column=0, sticky="nsew")

        self.safestop_x_label = tk.Label(
            self.jobs_error_view,
            text="X",
            bg=self.RED,
            fg="#FFFFFF",
            font=self.FONT_COUNTER,
        )  # text="✖",
        self.safestop_x_label.pack(expand=True)

        # show normal view at startup
        self.jobs_normal_view.tkraise()

        # 'online'-status animation
        self._online_animation_after_id = None
        self._online_pulse_index = 0

        # 'working...'-status animation
        self._working_animation_after_id = None
        self._working_dots = 0

    def _build_body(self, bg_color, text_color):
        self.body = tk.Frame(self.root, bg=bg_color)
        self.body.grid(row=1, column=0, sticky="nsew")
        self.body.grid_rowconfigure(0, weight=1)
        self.body.grid_columnconfigure(0, weight=1)

        # body content
        log_and_scroll_container = tk.Frame(self.body, bg=bg_color)
        log_and_scroll_container.grid(row=0, column=0, sticky="nsew")
        log_and_scroll_container.grid_rowconfigure(0, weight=1)
        log_and_scroll_container.grid_columnconfigure(0, weight=1)

        # the right-hand side scrollbar
        scrollbar = tk.Scrollbar(
            log_and_scroll_container,
            width=self.SCROLLBAR_WIDTH,
            troughcolor=self.SCROLL_TROUGH,
            bg=self.SCROLL_BG,
            activebackground=self.SCROLL_ACTIVE,
            bd=0,
            highlightthickness=0,
            relief="flat",
        )
        scrollbar.grid(row=0, column=1, sticky="ns")

        # the 'console'-style log
        self.log_text = tk.Text(
            log_and_scroll_container,
            yscrollcommand=scrollbar.set,
            bg=bg_color,
            fg=text_color,
            insertbackground="black",
            font=self.FONT_LOG,
            wrap="none",
            state="disabled",
            bd=0,
            highlightthickness=0,
        )  # glow highlightbackground="#1F2937", highlightthickness=1
        self.log_text.grid(row=0, column=0, sticky="nsew")
        scrollbar.config(command=self.log_text.yview)

    def _build_footer(self, bg_color, text_color):
        self.footer = tk.Frame(self.root, bg=bg_color)
        self.footer.grid(row=2, column=0, sticky="nsew")
        self.footer.grid_rowconfigure(0, weight=1)
        self.footer.grid_columnconfigure(0, weight=1)

        # footer content
        self.last_activity_label = tk.Label(
            self.footer,
            text="last activity: xx:xx",
            fg=self.MUTED,
            bg=bg_color,
            font=self.FONT_SMALL,
            anchor="e",
        )
        self.last_activity_label.grid(row=0, column=1, padx=8, pady=16)

    def _apply_status_update(self, status: DashboardStatus | None = None):

        # stops any ongoing animations
        self._stop_online_animation()
        self._stop_working_animation()
        self.status_dot.config(text="")

        # changes text
        if status == "online":
            self.rpa_status_label.config(text="online", fg=self.GREEN)
            self.jobs_normal_view.tkraise()
            self.status_dot.config(text="●")
            self._start_online_animation()

        elif status == "no_network":
            self.rpa_status_label.config(text="no network", fg=self.RED)
            self.jobs_normal_view.tkraise()

        elif status == "working":
            self.rpa_status_label.config(text="working...", fg=self.YELLOW)
            self.jobs_normal_view.tkraise()
            self._start_working_animation()

        elif status == "safestop":
            self.rpa_status_label.config(text="safestop", fg=self.RED)
            self.jobs_error_view.tkraise()

        elif status == "out_of_office":
            self.rpa_status_label.config(text="out of office", fg=self.YELLOW)
            self.jobs_normal_view.tkraise()

    def _apply_jobs_done_today(self, n) -> None:
        self.jobs_done_label.config(text=str(n))

    def _create_recording_overlay(self) -> None:
        # written by AI chatgpt 5.3 instant
        self.recording_win = tk.Toplevel(self.root)
        self.recording_win.withdraw()                # hidden at start
        self.recording_win.overrideredirect(True)    # no title/border
        self.recording_win.configure(bg="black")

        try:
            self.recording_win.attributes("-topmost", True)
        except Exception:
            pass

        width = self.RECORDING_WIDTH
        height = self.RECORDING_HEIGHT
        x = self.root.winfo_screenwidth() - width - self.RECORDING_MARGIN_RIGHT
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.recording_win.geometry(f"{width}x{height}+{x}+{y}")

        frame = tk.Frame(
            self.recording_win,
            bg="black",
            highlightbackground="#444444",
            highlightthickness=1,
            bd=0,
        )
        frame.pack(fill="both", expand=True)

        canvas = tk.Canvas(
            frame,
            width=44,
            height=44,
            bg="black",
            highlightthickness=0,
            bd=0,
        )
        canvas.place(x=18, y=33)
        canvas.create_oval(4, 4, 40, 40, fill=self.RED, outline=self.RED)

        label = tk.Label(
            frame,
            text="RECORDING",
            fg="#FFFFFF",
            bg="black",
            font=self.FONT_RECORDING,
            anchor="w",
        )
        label.place(x=75, y=33)

    def _show_recording_overlay(self) -> None:
        # written by AI chatgpt 5.3 instant
        try:
            width = self.RECORDING_WIDTH
            height = self.RECORDING_HEIGHT
            x = self.root.winfo_screenwidth() - width - self.RECORDING_MARGIN_RIGHT
            y = (self.root.winfo_screenheight() // 2) - (height // 2)
            self.recording_win.geometry(f"{width}x{height}+{x}+{y}")

            self.recording_win.deiconify()
            self.recording_win.lift()

            try:
                self.recording_win.attributes("-topmost", True)
            except Exception:
                pass
        except Exception:
            pass

    def _hide_recording_overlay(self) -> None:
        # hides recording window
        try:
            self.recording_win.withdraw()
        except Exception:
            pass

    def _start_working_animation(self):
        if self._working_animation_after_id is None:
            self._animate_working()

    def _animate_working(self):
        # written by AI chatgpt 5.3 instant
        states = ["working", "working.", "working..", "working..."]
        self._working_dots = (self._working_dots + 1) % len(states)
        self.rpa_status_label.config(text=states[self._working_dots])
        self._working_animation_after_id = self.root.after(500, self._animate_working)

    def _stop_working_animation(self):
        if self._working_animation_after_id is not None:
            self.root.after_cancel(self._working_animation_after_id)
            self._working_animation_after_id = None
            self._working_dots = 0

    def _start_online_animation(self):
        if self._online_animation_after_id is None:
            self._online_pulse_index = 0
            self._animate_online()

    def _animate_online(self):
        # green pulse animation
        colors = [self.GREEN, self.GREEN_2, self.BG, self.GREEN_3, self.GREEN_2]
        color = colors[self._online_pulse_index]

        self.status_dot.config(fg=color)

        self._online_pulse_index = (self._online_pulse_index + 1) % len(colors)
        self._online_animation_after_id = self.root.after(1000, self._animate_online)

    def _stop_online_animation(self):
        if self._online_animation_after_id is not None:
            self.root.after_cancel(self._online_animation_after_id)
            self._online_animation_after_id = None

    def _append_ui_log(self, log_line: str, blank_line_before: bool = False) -> None:

        self.log_text.config(state="normal")  # open for edit
        now = datetime.datetime.now().strftime("%H:%M")

        if blank_line_before:
            self.log_text.insert("end", "\n")

        self.log_text.insert("end", f"[{now}] {log_line}\n")

        self.log_text.config(state="disabled")  # closing edit
        self.log_text.see("end")

    def _on_close_attempt(self):
        message="Use STOP-button in RPA tool next time (or press 2 in rpa_tool_simulator.py)"
        self.post_log_line(message, blank_line_before=True)
        print(message)
        self.post_shutdown(delay=2000)

    def post_status_update(self, status: DashboardStatus) -> None:
        self.root.after(0, lambda: self._apply_status_update(status))

    def post_log_line(self, text: str, blank_line_before: bool = False) -> None:
        self.root.after(0, lambda: self._append_ui_log(text, blank_line_before))

    def post_show_recording_overlay(self) -> None:
        self.root.after(0, self._show_recording_overlay)

    def post_hide_recording_overlay(self) -> None:
        self.root.after(0, self._hide_recording_overlay)

    def post_jobs_done_today(self, n: int) -> None:
        self.root.after(0, lambda: self._apply_jobs_done_today(n))

    def post_shutdown(self, delay=0) -> None:
        self.root.after(delay, self.shutdown)


# ============================================================
# MAIN ENTRYPOINT
# ============================================================

class RobotRuntime:
    """Main orchestration runtime."""

    def __init__(self, ui, config):
        self.ui = ui
        self.config = config

        self.next_queryflow_check_time = 0
        self.prev_state: HandoverState | None = None
        self.rpa_tool_claim_started_at: float | None = None
        self.rpa_tool_execution_started_at: float | None = None

        self.logger = LoggerService(self.ui, config.system_log_path)
        self.handover = HandoverRepository(self.logger, config.handover_file)
        self.audit = AuditRepository(self.logger, config.audit_db_path)
        self.network_service = NetworkService(self.logger, config.network_healthcheck_path)
        self.recording = RecordingService(self.logger, config.recordings_in_progress_folder, config.recordings_destination_folder)
        
        
        if build_backends:
            backends = build_backends(self.logger)
            self.personal_mailbox = backends["personal_mailbox"]
            self.shared_mailbox = backends["shared_mailbox"]
            self.erp_backend = backends["erp_backend"]
        else:
            self.personal_mailbox = DemoMailBackend(self.logger, "personal_inbox")
            self.shared_mailbox = DemoMailBackend(self.logger, "shared_inbox")
            self.erp_backend = DemoErpBackend()


        self.personal_mail_handlers = {"ping": PingHandler(self.logger),}
        self.shared_mail_handlers = {}
        self.query_handlers = {}

        if build_custom_query_handlers is not None:
            self.query_handlers.update(build_custom_query_handlers(self.logger, self.audit, self.erp_backend,))

        if build_custom_personal_mail_handlers is not None:
            self.personal_mail_handlers.update(build_custom_personal_mail_handlers(self.logger,))

        if build_custom_shared_mail_handlers is not None:
            self.shared_mail_handlers.update(build_custom_shared_mail_handlers(self.logger,))
        
        self.job_handlers = {
            **self.personal_mail_handlers,
            **self.shared_mail_handlers,
            **self.query_handlers,
        }

        self._validate_job_handlers_registry()
        
        self.friends_repo = FriendsRepository(config.friends_path, allowed_job_names=set(self.personal_mail_handlers.keys()))
        self.notifications = UserNotificationService(self.personal_mailbox, config.recordings_destination_folder, config.rpa_tool_claim_timeout, config.rpa_tool_execution_timeout, config.rpa_admin_email,)
        self.job_lifecycle = JobLifecycleService(self.logger, self.handover, self.ui.post_show_recording_overlay, self.recording, self.audit, self.notifications, self.personal_mailbox, self.shared_mailbox, self.job_handlers, self.ui.post_hide_recording_overlay, self.generate_job_id)
        self.mail_flow = MailFlow(self.logger, self.friends_repo, self.audit, self._is_within_operating_hours, self.network_service, self.personal_mail_handlers, self.shared_mail_handlers, self.job_lifecycle, self.personal_mailbox, self.shared_mailbox)
        self.query_flow = QueryFlow(self.logger, self.query_handlers, self.job_lifecycle, self._is_within_operating_hours)
        self.mail_recovery = MailRecoveryService(self.logger, self.personal_mailbox, self.shared_mailbox, self.audit, self.friends_repo, self.notifications, self.generate_job_id) 
        self.safestop_controller = SafestopController(self.logger, self.recording, self.ui.post_hide_recording_overlay, self.ui.post_status_update, self.ui.post_shutdown, self._check_for_stop_flag, config.handover_file, self.mail_recovery, self.notifications) 
       
    def runtime_loop(self) -> None:
        handover_file: HandoverFile | None = None

        try:
            self._startup_sequence()
            
            while True:
                self._check_for_stop_flag()

                handover_file = self.handover.read()
                self._handle_state_change(handover_file)
                self._enforce_timeouts(handover_file)
                
                # Dispatch
                if handover_file.state == "idle":                 # RobotRuntime owns the workflow
                    self._poll_job_intake()

                elif handover_file.state == "job_queued":         # RPA Tool owns the workflow
                    pass

                elif handover_file.state == "job_running":        # RPA Tool owns the workflow
                    pass

                elif handover_file.state == "job_verifying":      # RobotRuntime owns the workflow
                    self.job_lifecycle.complete_from_handover(handover_file)
                    self._refresh_jobs_done_counter()
                    self.handover.write(HandoverFile(state="idle"))

                elif handover_file.state == "safestop":           # RobotRuntime owns the workflow
                    raise RpaToolCrash("unexpected stop", job_id=handover_file.job_id, handover_file=handover_file,) 

                time.sleep(self.config.poll_interval)


        except RuntimeFault as fault:
            fault.traceback_text = traceback.format_exc()
            self.safestop_controller.run_degraded_mode(fault)

        except Exception as err:
            fault = RuntimeFault(
                message=str(err),
                job_id=handover_file.job_id if handover_file else None,
                handover_file=handover_file,
                cause=err,
                traceback_text=traceback.format_exc(),
            )
            self.safestop_controller.run_degraded_mode(fault)
      
    def generate_job_id(self) -> int:
        ''' unique id for all jobs. This works under single-runtime-single-machine assumption'''

        job_id = int(datetime.datetime.now().strftime("%Y%m%d%H%M%S"))

        try:
            last_job_id = self.audit.get_latest_job_id()
        except Exception as e:
            time.sleep(5)
            last_job_id = int(datetime.datetime.now().strftime("%Y%m%d%H%M%S"))
            self.logger.system(f"WARN: using fallback last_job_id due to audit error={e}")

        job_id = max(job_id, last_job_id + 1)

        self.logger.system(f"assigned job_id", job_id)
        return job_id

    def _startup_sequence(self):
        handover_file: HandoverFile | None = None

        try:
            self.logger.system(f"RobotRuntime started, version={VERSION}, pid={os.getpid()}")

            # cleanup
            for f in ["stop.flag", "restart.flag"]:
                try: os.remove(f)
                except Exception: pass
       
            # to avoid unnecessary errors in demo mode (the intended way is RPA Tool creates handover.json)
            if not os.path.exists(self.handover.handover_file):
                self.handover.write(HandoverFile(state="idle"))

            handover_file = self.handover.read()
            if handover_file.state != "idle":
                raise PreHandoverCrash(
                    f"Expected {self.handover.handover_file} to start in idle, got {handover_file.state}",
                    handover_file=handover_file,
                )        

            atexit.register(self.recording.stop)
    
            self.audit.ensure_db_exists()
            self.network_service.has_network_access()
            self.recording.stop() # stop any remaining active recordings
            self.recording.cleanup_aborted_recordings()
            self.friends_repo.reload_if_modified()
            self._refresh_jobs_done_counter()
            
            self.mail_recovery.recover_pending_personal_mail_replies(reply_context="startup_recovery")
            self.mail_recovery.recover_stuck_shared_mail()
            

        except Exception as e:
            raise PreHandoverCrash(
                f"_startup_sequence failed: {e}",
                handover_file=handover_file,
                cause=e,
            ) from e

    def _refresh_jobs_done_counter(self, job_id=None):
        try:
            count = self.audit.count_done_jobs_today()
            self.ui.post_jobs_done_today(count)
        except Exception as err:
            self.logger.system(err, job_id)

    def _handle_state_change(self, handover_file: HandoverFile) -> None:
        '''do side effects on state change'''

        job_id = handover_file.job_id
        state = handover_file.state

        if state == self.prev_state:
            return

        transition_message=f"state transition detected by CPU-poll: {self.prev_state} -> {state}"

        if not self.handover.is_valid_observed_transition(self.prev_state, state):
            raise RuntimeError(f"invalid {transition_message}")

        self._update_dashboard_status(state)
        self.logger.system(transition_message, job_id)

        if state == "job_running":
            self.audit.update(job_id=job_id, lifecycle_status="RUNNING")
        
        self.prev_state = state

    def _enforce_timeouts(self, handover_file):
        state = handover_file.state
        now = time.time()

        if state == "job_queued":
            if self.rpa_tool_claim_started_at is None:
                self.rpa_tool_claim_started_at = now
                self.rpa_tool_execution_started_at = None
                return

            if now - self.rpa_tool_claim_started_at > self.config.rpa_tool_claim_timeout:
                # PreHandoverCrash signals to user that the robot has not started the job yet (design decision: RpaToolCrash?).
                raise PreHandoverCrash(
                    f"timeout before start (the request itself was correct)",
                    job_id=handover_file.job_id,
                    handover_file=handover_file,
                )

        elif state == "job_running":
            if self.rpa_tool_execution_started_at is None:
                self.rpa_tool_claim_started_at = None
                self.rpa_tool_execution_started_at = now
                return
        
            if now - self.rpa_tool_execution_started_at > self.config.rpa_tool_execution_timeout:
                raise RpaToolCrash(
                    f"timeout while working (the request itself was correct)",
                    job_id=handover_file.job_id,
                    handover_file=handover_file,
                )
        else:
            self.rpa_tool_claim_started_at = None
            self.rpa_tool_execution_started_at = None

    def _update_dashboard_status(self, state=None) -> None:
               
        if state is not None and state not in get_args(HandoverState):
            raise ValueError(f"unknown state: {state}")

        if state == "safestop":
            dashboard_status = "safestop"

        elif state in ("job_queued", "job_running", "job_verifying"):
            dashboard_status = "working"

        elif self.network_service.network_state is False:
            dashboard_status = "no_network"

        elif not self._is_within_operating_hours():
            dashboard_status = "out_of_office"

        else:
            dashboard_status = "online"

        self.ui.post_status_update(dashboard_status)

    def _poll_job_intake(self) -> bool:
        ''' job intake logic '''
        try:
            
            # 1. Mail first (priority)
            if self.mail_flow.poll_once():                
                return True
            
            # 2. Query (or other scheduled) jobs
            now = time.time()
            if now > self.next_queryflow_check_time:
                if self.query_flow.poll_once():
                    return True
                
                # set next check if no findings
                self.next_queryflow_check_time = now + self.config.queryflow_poll_interval 
            
            return False

        except PreHandoverCrash:
            raise
        except Exception as e:
            raise PreHandoverCrash(str(e), cause=e) from e

    def _is_within_operating_hours(self) -> bool:
        start = datetime.time(self.config.operating_hours_start)
        now = datetime.datetime.now().time()
        end = datetime.time(self.config.operating_hours_end) 
        return start <= now <= end

    def _validate_job_handlers_registry(self) -> None:
        # will not catch dublicate names for job handlers
        for key, handler in self.job_handlers.items():
            if not isinstance(key, str) or not key.strip():
                raise ValueError(f"invalid handler key: {key}")

            if not hasattr(handler, "job_name"):
                raise ValueError(f"handler {handler} missing job_name")

            if handler.job_name != key:
                raise ValueError(
                    f"handler registry mismatch: key={key}, handler.job_name={handler.job_name}"
                )

    def _check_for_stop_flag(self):
        ''' to stop main.py on operator manual stop on RPA tool '''

        stopflag = "stop.flag"
 
        if os.path.isfile(stopflag):
            try: os.remove(stopflag)
            except Exception: pass

            self.logger.system(f"found {stopflag}, initiating shutdown sequence")
            
            try: self.ui.post_shutdown() # request soft exit
            except Exception: os._exit(1)
            
            time.sleep(3)
            os._exit(0)  # kill if still alive after 3 sec 


def main() -> None:
    '''run UI in main thread and the rest async'''
    config = load_or_create_config()

    ui = DashboardUI()
    robot_runtime = RobotRuntime(ui, config)

    threading.Thread(target=robot_runtime.runtime_loop, daemon=True).start()
    ui.run()


if __name__ == "__main__":
    main()