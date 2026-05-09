# place in main.py directory

from __future__ import annotations
import random, os, time, uuid
from pathlib import Path
from email.message import EmailMessage
from email.utils import formatdate, make_msgid
from openpyxl import load_workbook  # type: ignore


class FakeEmailJobsGenerator:
    """to create fake email jobs"""

    BASE_DIR = Path(__file__).resolve().parent

    PERSONAL_PIPELINE_DIR = BASE_DIR / "personal_inbox"
    PERSONAL_INBOX_DIR = PERSONAL_PIPELINE_DIR / "inbox"
    SHARED_PIPELINE_DIR = BASE_DIR / "shared_inbox"
    SHARED_INBOX_DIR = SHARED_PIPELINE_DIR / "inbox"

    for folder in [
        PERSONAL_PIPELINE_DIR,
        PERSONAL_INBOX_DIR,
        SHARED_PIPELINE_DIR,
        SHARED_INBOX_DIR,
    ]:
        folder.mkdir(exist_ok=True)


    def build_email_message(
        self,
        *,
        from_name: str,
        from_email: str,
        to_email: str,
        subject: str,
        body: str,
        attachment_paths: list[Path] | None = None,
    ) -> EmailMessage:
        msg = EmailMessage()
        msg["From"] = f"{from_name} <{from_email}>"
        msg["To"] = to_email
        msg["Subject"] = subject
        msg["Date"] = formatdate(localtime=True)
        msg["Message-ID"] = make_msgid()
        msg.set_content(body)

        for path in attachment_paths or []:
            data = path.read_bytes()
            msg.add_attachment(
                data,
                maintype="application",
                subtype="octet-stream",
                filename=path.name,
            )

        return msg

    def write_eml(self, msg: EmailMessage, inbox_dir: Path, prefix: str = "mail") -> Path:
        """Atomic write into chosen inbox to reduce risk of partial reads."""
        unique_id = uuid.uuid4().hex[:12]
        final_path = inbox_dir / f"{prefix}_{unique_id}.eml"
        temp_path = inbox_dir / f".tmp_{prefix}_{unique_id}.eml"

        with open(temp_path, "wb") as f:
            f.write(msg.as_bytes())

        temp_path.replace(final_path)
        return final_path

    # -------------------------
    # personal inbox examples
    # -------------------------

    def create_ping_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Alice Wonderland",
            from_email="alice@example.com",
            to_email="robot@company.local",
            subject="PING",
            body=(
                "Hello,\n\n"
                "I'm sending you a ping\n"
                "BR,\n"
                "Alice"
            ),
        )
        return self.write_eml(msg, self.PERSONAL_INBOX_DIR, prefix="ping")

    def create_no_access_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Alice Wonderland",
            from_email="alice@example.com",
            to_email="robot@company.local",
            subject="Please run qty_adjust",
            body=(
                "I have no idea what qty_adjust is though...\n"
                "Best regards,\n"
                "Alice\n"
            ),
            #attachment_paths=[self.ATTACHMENTS_DIR / "qty_adjust_request.txt"],
        )
        return self.write_eml(msg, self.PERSONAL_INBOX_DIR, prefix="no_access")

    def create_valid_qty_adjust_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Bob Tester",
            from_email="bob@test.com",
            to_email="robot@company.local",
            subject="qty_adjust",
            body=(
                "Hello,\n\n"
                "Please run qty_adjust\n\n"
                "order_number: 100245\n"
                "order_qty: 12000\n"
                "material_available: 11031\n\n"
                "Best regards,\n"
                "Bob\n"
            ),
            #attachment_paths=[self.ATTACHMENTS_DIR / "qty_adjust_request.txt"],
        )
        return self.write_eml(msg, self.PERSONAL_INBOX_DIR, prefix="qty_adjust")


    def create_blocked_sender_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Mallory Intruder",
            from_email="mallory@evil.com",
            to_email="robot@company.local",
            subject="Please run qty_adjust",
            body=(
                "Hello,\n\n"
                "I would like the robot to run qty_adjust.\n\n"
                "Regards,\n"
                "Mallory\n"
            ),
        )
        return self.write_eml(msg, self.PERSONAL_INBOX_DIR, prefix="blocked")

    # -------------------------
    # shared inbox examples
    # -------------------------

    def create_shared_supplier1_order_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Supplier One",
            from_email="supplier1@example.com",
            to_email="orderconfirmations@company.local",
            subject="Order confirmation SO-100245",
            body=(
                "Hello,\n\n"
                "Please find order confirmation for order 100245 below.\n\n"
                "order_number: 100245\n"
                "confirmed_qty: 12000\n"
                "eta: 2027-11-25\n\n"
                "Best regards,\n"
                "Supplier One\n"
            ),
        )
        return self.write_eml(msg, self.SHARED_INBOX_DIR, prefix="shared")

    def create_faulty_shared_supplier1_order_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Supplier One",
            from_email="supplier1@example.com",
            to_email="orderconfirmations@company.local",
            subject="Order confirmation SO-100246",
            body=(
                "Hello,\n\n"
                "Please find order confirmation for order 100246 below.\n\n"
                "order_number: 100246\n"
                "confirmed_qty: -11\n" # the error
                "eta: 2027-11-26\n\n"
                "Best regards,\n"
                "Supplier One\n"
            ),
        )
        return self.write_eml(msg, self.SHARED_INBOX_DIR, prefix="shared")

    def create_shared_outofscope_supplier_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="New Supplier",
            from_email="new_supplier@example.com",
            to_email="orderconfirmations@company.local",
            subject="Order confirmation 100477",
            body=(
                "Hello,\n\n"
                "Order confirmation attached in body text only.\n\n"
                "order_number: 100477\n"
                "confirmed_qty: 5000\n"
                "eta: 2027-12-04\n\n"
                "Kind regards,\n"
                "New Supplier\n"
            ),
        )
        return self.write_eml(msg, self.SHARED_INBOX_DIR, prefix="out_of_scope")


class FakeQueryJobsGenerator:
    """to create fake ERP jobs"""

    def add_random_row(self, path="Demo_ERP_table.xlsx") -> str:
        """this example will be classified as order_adjust in RobotRuntime"""

        if not os.path.isfile(path):
            raise RuntimeError(f"{path} not found, run main.py first")
        wb = load_workbook(path)
        ws = wb.active

        assert ws is not None

        next_row = ws.max_row + 1

        erp_order_number = str(random.randint(10000000, 10999999))
        order_qty = random.randint(10, 100) * 100
        material_available = order_qty + random.randint(-100, 100)

        ws[f"A{next_row}"] = erp_order_number
        ws[f"B{next_row}"] = order_qty
        ws[f"C{next_row}"] = material_available

        wb.save(path)
        wb.close()
        return erp_order_number


class FakeJobsGenerator:
    """produce a fake email or a fake query-job at random"""

    def __init__(self) -> None:
        self.fake_emailjob = FakeEmailJobsGenerator()
        self.fake_queryjob = FakeQueryJobsGenerator()

    def run(self):
        while True:
            try:
                input("\nHit Enter to generate a random job")

                creators = [
                    # personal inbox
                    self.fake_emailjob.create_ping_mail,
                    self.fake_emailjob.create_valid_qty_adjust_mail,
                    self.fake_emailjob.create_no_access_mail,
                    self.fake_emailjob.create_blocked_sender_mail,

                    # shared inbox
                    self.fake_emailjob.create_shared_supplier1_order_mail,
                    self.fake_emailjob.create_faulty_shared_supplier1_order_mail,
                    self.fake_emailjob.create_shared_outofscope_supplier_mail,

                    # query
                    self.fake_queryjob.add_random_row
                ]

                created = random.choice(creators)()
                if isinstance(created, Path):
                    print(f"Created email job: {created.name}")
                else:
                    print(f"Created query job: order_number={created}")


            except KeyboardInterrupt:
                print("\nStopped.")
                break
            except Exception as err:
                print(f"WARN: generator error: {err}")
                time.sleep(1)


def main():
    if not os.path.isfile("main.py"):
        raise RuntimeError("Place this file in main.py directory")

    FakeJobsGenerator().run()


if __name__ == "__main__":
    main()